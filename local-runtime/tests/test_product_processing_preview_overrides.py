from __future__ import annotations

from pathlib import Path

from wh_local.modules.product_processing.domain.workbooks import _dxm_single_export_row
from wh_local.modules.product_processing.infrastructure.assets import ProductProcessingAssets
from wh_local.modules.product_processing.infrastructure.database import create_database
from wh_local.modules.product_processing.infrastructure.repository import (
    PreviewSlotConflict,
    ProductProcessingRepository,
)
from wh_local.modules.product_processing.service import ProductProcessingService


def _service(tmp_path: Path) -> ProductProcessingService:
    return ProductProcessingService(
        ProductProcessingRepository(create_database("sqlite:///:memory:")),
        ProductProcessingAssets(tmp_path / "assets"),
    )


def _base_result() -> dict:
    return {
        "product_draft_id": 1,
        "skc": "SKC-1",
        "sku": "SKU-1",
        "category": "Home & Kitchen",
        "category_path": "Home & Kitchen > Drinkware",
        "category_id": "12345",
        "optimized_title": "Original AI Generated Title",
        "description": "DURABLE MATERIAL - Made of stainless steel.",
        "image_url": "https://src.example.com/main.jpg",
        "source_url": "https://src.example.com/product",
        "source_image_urls": ["https://src.example.com/1.jpg", "https://src.example.com/2.jpg"],
        "source_detail_image_urls": ["https://src.example.com/d1.jpg"],
        "source_attributes": [],
        "source_variant_records": [],
        "variant_value_translations": {},
        "cost": 40.0,
        "declared_price": 160.0,
        "suggested_price": 40.0,
        "product_dimensions": {"length_cm": 20, "width_cm": 15, "height_cm": 10, "weight_g": 300},
        "stock": 50,
        "carousel_image_paths": [
            "https://cos.example.com/c1.jpg",
            "https://cos.example.com/c2.jpg",
            "https://cos.example.com/c3.jpg",
            "https://cos.example.com/c4.jpg",
        ],
        "grid_image_summary_path": "https://cos.example.com/summary.jpg",
        "detail_image_paths": ["https://cos.example.com/detail.jpg"],
        "status": "completed",
        "preflight_only": False,
    }


def _create_task_with_result(service: ProductProcessingService) -> dict:
    draft, _ = service.create_draft(
        {"source_type": "manual", "title": "Source Title", "product_name": "Source Title", "skc": "SKC-1"},
        workspace_id="local",
    )
    result = _base_result()
    result["product_draft_id"] = draft["id"]
    task = service.repository.create_task(
        title="预检测试",
        preflight_only=False,
        settings={"target_site": "US", "target_language": "en"},
        drafts=[draft],
        idempotency_key=None,
        workspace_id="local",
    )
    item = task["items"][0]
    finished = service.repository.finish_task(
        task["id"],
        [
            {
                "item_id": item["id"],
                "status": "completed",
                "reason": "",
                "title": result["optimized_title"],
                "image_url": result["image_url"],
                "result": result,
            }
        ],
        output_file=f"task_{task['id']}/dxm_import_task_{task['id']}.xlsx",
        error_report_file=f"task_{task['id']}/error_report_task_{task['id']}.csv",
        video_manifest_file="",
        workspace_id="local",
    )
    return finished


def test_preview_default_matches_generated_results(tmp_path: Path) -> None:
    service = _service(tmp_path)
    task = _create_task_with_result(service)
    preview = service.task_preview(task["id"], workspace_id="local")
    assert preview["item_count"] == 1
    item = preview["items"][0]
    assert item["title"] == "Original AI Generated Title"
    assert item["source_url"] == "https://src.example.com/product"
    assert item["overrides"] == {}
    assert item["carousel_images"][0] == "https://cos.example.com/c1.jpg"
    assert item["main_image"] == "https://cos.example.com/c1.jpg"
    assert item["core_fields"]["declared_price"] == 160.0
    assert item["core_fields"]["length_cm"] == 20


def test_save_preview_overrides_then_preview_merges(tmp_path: Path) -> None:
    service = _service(tmp_path)
    task = _create_task_with_result(service)
    draft_id = task["items"][0]["product_draft_id"]
    revision = service.task_preview(task["id"], workspace_id="local")["items"][0]["preview_revision"]
    saved = service.save_task_preview(
        task["id"],
        [
            {
                "product_draft_id": draft_id,
                "expected_preview_revision": revision,
                "overrides": {
                    "title": "Manual Edited Title",
                    "carousel_images": ["https://user.example.com/new1.jpg"],
                    "core_fields": {"declared_price": 888, "stock": 7},
                },
            }
        ],
        workspace_id="local",
    )
    assert saved["saved_count"] == 1
    preview = service.task_preview(task["id"], workspace_id="local")
    item = preview["items"][0]
    assert item["title"] == "Manual Edited Title"
    assert item["carousel_images"] == ["https://user.example.com/new1.jpg"]
    assert item["main_image"] == "https://user.example.com/new1.jpg"
    assert item["core_fields"]["declared_price"] == 888
    assert item["core_fields"]["stock"] == 7


def test_clean_preview_overrides_drops_empty_values() -> None:
    cleaned = ProductProcessingService._clean_preview_overrides(
        {
            "title": "",
            "description": "  ",
            "main_image": "https://x.example.com/m.jpg",
            "carousel_images": [],
            "detail_images": ["https://x.example.com/d.jpg", ""],
            "image_slot_overrides": {
                "carousel.dimension_background": {"url": "https://x.example.com/dimension.jpg"},
                "carousel.detail": {"url": ""},
            },
            "core_fields": {"sku": "", "declared_price": None, "stock": 5, "category_path": "  "},
        }
    )
    assert cleaned == {
        "main_image": "https://x.example.com/m.jpg",
        "detail_images": ["https://x.example.com/d.jpg"],
        "image_slot_overrides": {
            "carousel.dimension_background": {"url": "https://x.example.com/dimension.jpg"},
        },
        "core_fields": {"stock": 5},
    }


def test_preview_revision_changes_only_when_overrides_change(tmp_path: Path) -> None:
    service = _service(tmp_path)
    task = _create_task_with_result(service)
    draft_id = task["items"][0]["product_draft_id"]
    overrides = {"title": "Manual Edited Title"}

    first = service.repository.save_draft_preview_overrides(draft_id, overrides)
    unchanged = service.repository.save_draft_preview_overrides(draft_id, overrides)

    assert first is not None
    assert unchanged is not None
    assert first["preview_revision"] == 1
    assert unchanged["preview_revision"] == 1


def test_dimension_accept_preserves_unrelated_preview_edits(tmp_path: Path) -> None:
    service = _service(tmp_path)
    task = _create_task_with_result(service)
    draft_id = task["items"][0]["product_draft_id"]
    service.repository.save_draft_preview_overrides(
        draft_id,
        {"title": "Edited after canvas import"},
    )

    updated = service.repository.apply_dimension_slot_patch(
        draft_id,
        target_slot="carousel.dimension_background",
        patch={"url": "https://user.example.com/dimension.jpg", "asset_id": "asset-1"},
        base_slot_value="https://cos.example.com/c4.jpg",
    )

    assert updated is not None
    assert updated["preview_revision"] == 2
    assert updated["preview_overrides"]["title"] == "Edited after canvas import"
    assert updated["preview_overrides"]["image_slot_overrides"] == {
        "carousel.dimension_background": {
            "url": "https://user.example.com/dimension.jpg",
            "asset_id": "asset-1",
        }
    }


def test_dimension_accept_rejects_newer_target_slot_edit(tmp_path: Path) -> None:
    service = _service(tmp_path)
    task = _create_task_with_result(service)
    draft_id = task["items"][0]["product_draft_id"]
    service.repository.save_draft_preview_overrides(
        draft_id,
        {
            "image_slot_overrides": {
                "carousel.dimension_background": {"url": "https://user.example.com/newer.jpg"}
            }
        },
    )

    import pytest

    with pytest.raises(PreviewSlotConflict):
        service.repository.apply_dimension_slot_patch(
            draft_id,
            target_slot="carousel.dimension_background",
            patch={"url": "https://user.example.com/stale-canvas.jpg"},
            base_slot_value="https://cos.example.com/c4.jpg",
        )


def test_dimension_slot_patch_preserves_other_carousel_and_summary(tmp_path: Path) -> None:
    service = _service(tmp_path)
    task = _create_task_with_result(service)
    draft_id = task["items"][0]["product_draft_id"]
    revision = service.task_preview(task["id"], workspace_id="local")["items"][0]["preview_revision"]
    service.save_task_preview(
        task["id"],
        [
            {
                "product_draft_id": draft_id,
                "expected_preview_revision": revision,
                "overrides": {
                    "image_slot_overrides": {
                        "carousel.dimension_background": {
                            "url": "https://user.example.com/dimension.jpg",
                            "asset_id": "dimension-asset-1",
                        }
                    }
                },
            }
        ],
        workspace_id="local",
    )

    preview = service.task_preview(task["id"], workspace_id="local")
    item = preview["items"][0]
    assert item["carousel_images"] == [
        "https://cos.example.com/c1.jpg",
        "https://cos.example.com/c2.jpg",
        "https://cos.example.com/c3.jpg",
        "https://user.example.com/dimension.jpg",
        "https://cos.example.com/summary.jpg",
    ]
    assert item["image_slots"][3]["slot_id"] == "carousel.dimension_background"

    exported = _dxm_single_export_row({**_base_result(), "preview_overrides": item["overrides"]}, None)
    assert exported[18].splitlines() == [
        "https://cos.example.com/c1.jpg",
        "https://cos.example.com/c2.jpg",
        "https://cos.example.com/c3.jpg",
        "https://user.example.com/dimension.jpg",
        "https://cos.example.com/summary.jpg",
    ]


def test_dimension_slot_patch_uses_legacy_carousel_as_its_baseline() -> None:
    row = _base_result()
    row["image_manifest"] = [
        {"slot_id": "carousel.hero", "role": "hero", "value": "https://cos.example.com/c1.jpg"},
        {"slot_id": "carousel.detail", "role": "detail", "value": "https://cos.example.com/c2.jpg"},
        {"slot_id": "carousel.lifestyle", "role": "lifestyle", "value": "https://cos.example.com/c3.jpg"},
        {
            "slot_id": "carousel.dimension_background",
            "role": "dimension_background",
            "value": "https://cos.example.com/c4.jpg",
        },
    ]
    row["preview_overrides"] = {
        "carousel_images": [
            "https://user.example.com/legacy1.jpg",
            "https://user.example.com/legacy2.jpg",
            "https://user.example.com/legacy3.jpg",
            "https://user.example.com/legacy4.jpg",
        ],
        "image_slot_overrides": {
            "carousel.dimension_background": {"url": "https://user.example.com/dimension.jpg"}
        },
    }

    exported = _dxm_single_export_row(row, None)

    assert exported[18].splitlines() == [
        "https://user.example.com/legacy1.jpg",
        "https://user.example.com/legacy2.jpg",
        "https://user.example.com/legacy3.jpg",
        "https://user.example.com/dimension.jpg",
        "https://cos.example.com/summary.jpg",
    ]


def test_export_final_workbook_applies_overrides(tmp_path: Path) -> None:
    service = _service(tmp_path)
    task = _create_task_with_result(service)
    draft_id = task["items"][0]["product_draft_id"]
    revision = service.task_preview(task["id"], workspace_id="local")["items"][0]["preview_revision"]
    service.save_task_preview(
        task["id"],
        [
            {
                "product_draft_id": draft_id,
                "expected_preview_revision": revision,
                "overrides": {
                    "title": "Manual Edited Title",
                    "main_image": "https://user.example.com/main.jpg",
                    "carousel_images": [
                        "https://user.example.com/c1.jpg",
                        "https://user.example.com/c2.jpg",
                    ],
                    "core_fields": {"declared_price": 999, "length_cm": 30, "weight_g": 500},
                },
            }
        ],
        workspace_id="local",
    )
    exported = service.export_final_workbook(task["id"], workspace_id="local")
    assert exported["row_count"] == 1
    path = service.assets.output_root / f"task_{task['id']}" / exported["file"]
    assert path.is_file()

    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    row = dict(zip(headers, rows[0]))
    assert row["*产品标题"] == "Manual Edited Title"
    assert row["*英文标题"] == "Manual Edited Title"
    assert row["预览图"] == "https://user.example.com/main.jpg"
    assert row["*产品素材图"] == "https://user.example.com/main.jpg"
    assert row["*轮播图"] == "https://user.example.com/c1.jpg\nhttps://user.example.com/c2.jpg"
    assert row["*申报价格\n(店铺币种)"] == 999
    assert row["*长（cm）"] == 30
    # 人工确认的实际重量必须原样导出，不能被体积重兜底静默改写。
    assert row["*重量（g）"] == 500


def test_dxm_single_export_row_defaults_without_overrides() -> None:
    row = _base_result()
    values = _dxm_single_export_row(row, None)
    assert values[0] == "Original AI Generated Title"
    assert values[8] == "https://cos.example.com/c1.jpg"
    assert values[18] == "\n".join(
        [
            "https://cos.example.com/c1.jpg",
            "https://cos.example.com/c2.jpg",
            "https://cos.example.com/c3.jpg",
            "https://cos.example.com/c4.jpg",
            "https://cos.example.com/summary.jpg",
        ]
    )
    # 系统生成值应用体积重兜底后，再按店小秘规则向上取整到 100g。
    assert values[14] == 600
    assert values[19] == "https://cos.example.com/c1.jpg"
