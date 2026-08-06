import tempfile
import unittest
from pathlib import Path

from fastapi.testclient import TestClient

from wh_local.app.main import create_app


class ProductProcessingApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
        self.db_path = Path(self.temp_dir.name) / "test_wh_local.sqlite3"
        self.app = create_app(database_path=self.db_path)
        self.client = TestClient(self.app)
        self.workspace = "test-workspace"
        self.headers = {"X-Workspace-ID": self.workspace}

    def tearDown(self) -> None:
        self.client.close()
        self.temp_dir.cleanup()

    def test_engine_status(self) -> None:
        response = self.client.get("/api/product-processing/engine/status", headers=self.headers)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["available"])
        self.assertIn("ready", data)
        self.assertIn("diagnostics", data)

    def test_demo_draft_and_list(self) -> None:
        create_response = self.client.post("/api/product-processing/demo-draft", headers=self.headers)
        self.assertEqual(create_response.status_code, 200)
        draft = create_response.json()["draft"]
        self.assertIn("id", draft)
        self.assertEqual(draft["workspace_id"], self.workspace)

        list_response = self.client.get(
            "/api/product-processing/drafts?view=summary&limit=10", headers=self.headers
        )
        self.assertEqual(list_response.status_code, 200)
        data = list_response.json()
        self.assertEqual(len(data["drafts"]), 1)
        self.assertEqual(data["drafts"][0]["id"], draft["id"])

    def test_process_draft_and_outputs(self) -> None:
        create_response = self.client.post("/api/product-processing/demo-draft", headers=self.headers)
        draft_id = create_response.json()["draft"]["id"]

        process_response = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "test-process-1"},
            json={
                "title": "测试处理任务",
                "draft_ids": [draft_id],
                "max_products": 10,
                "preflight_only": False,
                "target_site": "US",
                "target_language": "en",
            },
        )
        self.assertEqual(process_response.status_code, 200)
        task_id = process_response.json()["task_id"]
        self.assertIsInstance(task_id, int)

        outputs_response = self.client.get(
            f"/api/product-processing/tasks/{task_id}/outputs", headers=self.headers
        )
        self.assertEqual(outputs_response.status_code, 200)
        outputs = outputs_response.json()
        self.assertEqual(outputs["task_id"], task_id)
        self.assertEqual(outputs["total_count"], 1)
        self.assertGreaterEqual(outputs["success_count"], 0)
        self.assertIn("items", outputs)
        self.assertIn("outputs", outputs)

    def test_preflight_draft(self) -> None:
        create_response = self.client.post("/api/product-processing/demo-draft", headers=self.headers)
        draft_id = create_response.json()["draft"]["id"]

        response = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "test-preflight-1"},
            json={
                "title": "测试预检任务",
                "draft_ids": [draft_id],
                "preflight_only": True,
                "target_site": "US",
                "target_language": "en",
            },
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["task"]["metadata"]["preflight_only"])

    def test_download_dxm_output(self) -> None:
        create_response = self.client.post("/api/product-processing/demo-draft", headers=self.headers)
        draft_id = create_response.json()["draft"]["id"]

        process_response = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "test-download-1"},
            json={
                "title": "测试下载任务",
                "draft_ids": [draft_id],
                "target_site": "US",
                "target_language": "en",
            },
        )
        task_id = process_response.json()["task_id"]

        download_response = self.client.get(
            f"/api/product-processing/tasks/{task_id}/download?kind=dxm", headers=self.headers
        )
        self.assertEqual(download_response.status_code, 200)
        self.assertEqual(
            download_response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertGreater(len(download_response.content), 0)

    def test_process_draft_with_prototype_scope_options(self) -> None:
        create_response = self.client.post("/api/product-processing/demo-draft", headers=self.headers)
        draft_id = create_response.json()["draft"]["id"]

        process_response = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "test-scope-1"},
            json={
                "title": "测试处理任务-原型选项",
                "draft_ids": [draft_id],
                "max_products": 10,
                "target_site": "CO",
                "target_language": "es",
                "processing_scope": ["title", "details", "qualification"],
                "qualification_mode": "strict",
                "include_product_video": True,
            },
        )
        self.assertEqual(process_response.status_code, 200)
        data = process_response.json()
        task_id = data["task_id"]

        # 新请求体字段应被解析到任务响应中
        self.assertEqual(data["target_site"], "CO")
        self.assertEqual(data["target_language"], "es")
        self.assertEqual(data["qualification_mode"], "strict")
        self.assertTrue(data["include_product_video"])
        self.assertIn("title", data["processing_scope"])
        self.assertNotIn("four_grid", data["processing_scope"])

        # 任务元数据应保留归一化后的 settings
        self.assertEqual(data["task"]["metadata"]["settings"]["processing_scope"], ["title", "details", "qualification"])
        self.assertTrue(data["task"]["metadata"]["settings"]["product_video_template"])
        self.assertFalse(data["task"]["metadata"]["settings"]["grid_image"])

        outputs_response = self.client.get(
            f"/api/product-processing/tasks/{task_id}/outputs", headers=self.headers
        )
        self.assertEqual(outputs_response.status_code, 200)
        outputs = outputs_response.json()
        self.assertIn("manifest", outputs)
        self.assertIn("item_counts", outputs["manifest"])
        self.assertIn("artifacts", outputs)
        self.assertIn("attention_required_count", outputs)
        self.assertIn("technical_retryable_count", outputs)
        self.assertIn("configuration_blocked_count", outputs)
        self.assertIn("identity_review_required_count", outputs)
        self.assertIn("logistics_review_required_count", outputs)

    def test_legacy_boolean_options_still_work(self) -> None:
        create_response = self.client.post("/api/product-processing/demo-draft", headers=self.headers)
        draft_id = create_response.json()["draft"]["id"]

        process_response = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "test-legacy-1"},
            json={
                "title": "测试处理任务-旧版选项",
                "draft_ids": [draft_id],
                "title_optimize": False,
                "description": True,
                "grid_image": False,
                "target_site": "US",
                "target_language": "en",
            },
        )
        self.assertEqual(process_response.status_code, 200)
        data = process_response.json()
        scope = data["processing_scope"]
        self.assertNotIn("title", scope)
        self.assertIn("details", scope)
        self.assertNotIn("four_grid", scope)
        self.assertEqual(data["qualification_mode"], "standard")
        self.assertFalse(data["include_product_video"])

    def test_ai_config_endpoint(self) -> None:
        response = self.client.get("/api/product-processing/ai-config", headers=self.headers)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["provider"], "aicoming")
        self.assertEqual(data["base_url"], "https://api.aicoming.top/v1")
        self.assertTrue(data["api_key_configured"])
        masked = data["api_key_masked"]
        self.assertLess(len(masked), 40)
        self.assertTrue(masked.startswith("sk-") or masked.startswith("***"))
        self.assertNotIn("api_key", data)
        self.assertIn("text_model", data)
        self.assertIn("image_model", data)

    def test_skip_duplicates_skips_processed_drafts(self) -> None:
        create_response = self.client.post("/api/product-processing/demo-draft", headers=self.headers)
        draft_id = create_response.json()["draft"]["id"]

        first = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "skip-dup-1"},
            json={"draft_ids": [draft_id], "target_site": "US", "target_language": "en"},
        )
        self.assertEqual(first.status_code, 200)
        self.assertEqual(first.json()["success_count"], 1)

        # 第二次勾选同一草稿且开启“跳过已处理”，不应重复处理
        second = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "skip-dup-2"},
            json={
                "draft_ids": [draft_id],
                "target_site": "US",
                "target_language": "en",
                "skip_duplicates": True,
            },
        )
        self.assertEqual(second.status_code, 200)
        data = second.json()
        self.assertEqual(data["status"], "skipped")
        self.assertIn("跳过已处理", data["message"])

        # 未开启“跳过已处理”时仍可重新处理
        third = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "skip-dup-3"},
            json={"draft_ids": [draft_id], "target_site": "US", "target_language": "en"},
        )
        self.assertEqual(third.status_code, 200)
        self.assertEqual(third.json()["success_count"], 1)

    def test_ip_check_off_allows_risk_tagged_product(self) -> None:
        run_id = "ip-off-run-1"
        intake = self.client.post(
            "/api/product-processing/intake/daily-selection",
            headers=self.headers,
            json={
                "run_id": run_id,
                "workspace_id": self.workspace,
                "status": "completed",
                "criteria": {"keywords": ["保温杯"], "target_count": 1},
                "candidate_count": 1,
                "counts": {"total": 1, "confirmed": 1},
                "candidates": [
                    {
                        "candidate_id": f"{run_id}:c1",
                        "offer_id": "1688:9001",
                        "source_platform": "1688",
                        "source_url": "https://detail.1688.com/offer/9001.html",
                        "source_title": "品牌联名保温杯",
                        "main_image_url": "https://example.com/ip-off-c1.jpg",
                        "source_image_urls": ["https://example.com/ip-off-c1.jpg"],
                        "price_cny": "28.0",
                        "selection_score": "75",
                        "risk_tags": ["trademark"],
                    }
                ],
            },
        )
        self.assertEqual(intake.status_code, 200)
        draft_id = intake.json()["drafts"][0]["id"]

        # 侵权词过滤关闭时，带商标风险标签的商品应可处理成功
        process = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "ip-off-1"},
            json={
                "draft_ids": [draft_id],
                "target_site": "US",
                "target_language": "en",
                "ip_check": False,
            },
        )
        self.assertEqual(process.status_code, 200)
        self.assertEqual(process.json()["success_count"], 1)
        self.assertEqual(process.json()["failed_count"], 0)

        # 侵权词过滤开启时，同一草稿应被配置阻断
        process2 = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "ip-off-2"},
            json={
                "draft_ids": [draft_id],
                "target_site": "US",
                "target_language": "en",
                "ip_check": True,
            },
        )
        self.assertEqual(process2.status_code, 200)
        data = process2.json()
        self.assertEqual(data["attention_required_count"], 1)
        self.assertIn("侵权词过滤", data["items"][0]["reason"])

    def test_dxm_output_columns_match_prototype(self) -> None:
        create_response = self.client.post("/api/product-processing/demo-draft", headers=self.headers)
        draft_id = create_response.json()["draft"]["id"]

        process = self.client.post(
            "/api/product-processing/drafts/process",
            headers={**self.headers, "Idempotency-Key": "dxm-columns-1"},
            json={"draft_ids": [draft_id], "target_site": "US", "target_language": "en"},
        )
        self.assertEqual(process.status_code, 200)
        task_id = process.json()["task_id"]

        download = self.client.get(
            f"/api/product-processing/tasks/{task_id}/download?kind=dxm", headers=self.headers
        )
        self.assertEqual(download.status_code, 200)
        from io import BytesIO

        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(download.content), read_only=True)
        sheet = workbook.active
        headers = [str(cell) for cell in next(sheet.iter_rows(values_only=True))]
        workbook.close()
        for required in ("*产品标题", "*英文标题", "产品描述", "产品货号", "*变种属性名称一",
                         "*申报价格\n(店铺币种)", "SKU货号", "*轮播图", "*产品素材图",
                         "建议售价（USD）", "SKU分类"):
            self.assertIn(required, headers, f"DXM 导出缺少列: {required}")

    def test_invalid_processing_scope_rejected(self) -> None:
        response = self.client.post(
            "/api/product-processing/drafts/process",
            headers=self.headers,
            json={
                "draft_ids": [1],
                "processing_scope": ["not_a_real_scope"],
            },
        )
        self.assertEqual(response.status_code, 422)


if __name__ == "__main__":
    unittest.main()
