# -*- coding: utf-8 -*-
"""店小秘导出：变种属性值翻译与规格轴名称映射（对齐原型 VARIANT_VALUE_TRANSLATION_PROMPT / §8.2）。"""
from __future__ import annotations

import re
from pathlib import Path
from tempfile import TemporaryDirectory

from wh_local.modules.product_processing.domain.workbooks import create_result_workbook


def _chinese_in(value: object) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", str(value or "")))


def test_variant_values_use_translations_and_axis_names() -> None:
    row: dict = {
        "optimized_title": "Kitchen Faucet Extender",
        "description": "Desc",
        "skc": "SKC1",
        "sku": "SKU1",
        "cost": 20,
        "image_url": "http://img/main.jpg",
        "source_image_urls": ["http://img/1.jpg"],
        "source_detail_image_urls": [],
        "source_attributes": [],
        "source_variant_records": [
            {"sku_id": "S1", "attributes": {"规格": "【可旋转/白色盒装】"}},
            {"sku_id": "S2", "attributes": {"规格": "【更换滤芯/十片装】"}},
        ],
        "variant_value_translations": {
            "【可旋转/白色盒装】": "Rotatable white box pack",
            "【更换滤芯/十片装】": "Replacement filter cores, 10 pcs",
        },
        "product_dimensions": {"length_cm": 30, "width_cm": 16, "height_cm": 6, "weight_g": 160},
    }
    with TemporaryDirectory() as tmp:
        target = Path(tmp) / "dxm.xlsx"
        create_result_workbook([row], target)
        from openpyxl import load_workbook

        workbook = load_workbook(target, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
    assert len(rows) == 3  # 表头 + 2 个变种行
    first, second = rows[1], rows[2]
    # 规格轴名称映射为 Style，值用翻译
    assert first[4] == "Style"
    assert first[5] == "Rotatable white box pack"
    assert second[5] == "Replacement filter cores, 10 pcs"
    # 变种列（4-7 列）不得含中文（外包装等平台枚举列允许中文）
    assert not any(_chinese_in(value) for value in first[4:8])
    assert not any(_chinese_in(value) for value in second[4:8])


def test_operator_display_name_preferred_over_translation() -> None:
    row: dict = {
        "optimized_title": "Mesh Bag",
        "description": "Desc",
        "skc": "SKC1",
        "sku": "SKU1",
        "cost": 20,
        "image_url": "http://img/main.jpg",
        "source_image_urls": ["http://img/1.jpg"],
        "source_detail_image_urls": [],
        "source_attributes": [],
        "source_variant_records": [
            {"sku_id": "S1", "attributes": {"颜色": "米色"}, "display_name": "Diamond mesh bag - Beige"},
        ],
        "variant_value_translations": {"米色": "Beige"},
        "product_dimensions": {"length_cm": 32, "width_cm": 24, "height_cm": 5, "weight_g": 240},
    }
    with TemporaryDirectory() as tmp:
        target = Path(tmp) / "dxm.xlsx"
        create_result_workbook([row], target)
        from openpyxl import load_workbook

        workbook = load_workbook(target, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
    assert rows[1][4] == "Color"
    assert rows[1][5] == "Diamond mesh bag - Beige"


def test_no_translation_table_falls_back_to_raw_value() -> None:
    row: dict = {
        "optimized_title": "Faucet Extender",
        "description": "Desc",
        "skc": "SKC1",
        "sku": "SKU1",
        "cost": 20,
        "image_url": "http://img/main.jpg",
        "source_image_urls": ["http://img/1.jpg"],
        "source_detail_image_urls": [],
        "source_attributes": [],
        "source_variant_records": [
            {"sku_id": "S1", "attributes": {"规格": "【白色盒装】"}},
        ],
        "product_dimensions": {"length_cm": 30, "width_cm": 16, "height_cm": 6, "weight_g": 160},
    }
    with TemporaryDirectory() as tmp:
        target = Path(tmp) / "dxm.xlsx"
        create_result_workbook([row], target)
        from openpyxl import load_workbook

        workbook = load_workbook(target, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
    assert rows[1][5] == "【白色盒装】"
