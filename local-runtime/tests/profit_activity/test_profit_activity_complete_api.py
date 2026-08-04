"""Regression tests for the complete local Profit Activity screen contract."""

from __future__ import annotations

import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


LOCAL_RUNTIME = Path(__file__).resolve().parents[2]
if str(LOCAL_RUNTIME) not in sys.path:
    sys.path.insert(0, str(LOCAL_RUNTIME))

from wh_local.modules.profit_activity import create_profit_activity_router, create_profit_activity_service


class ProfitActivityCompleteApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        root = Path(self.temp.name)
        self.service = create_profit_activity_service(f"sqlite:///{(root / 'profit_activity.sqlite3').as_posix()}")
        app = FastAPI()
        app.include_router(create_profit_activity_router(self.service))
        app.include_router(create_profit_activity_router(self.service), prefix="/api/v1")
        self.client = TestClient(app)
        self.output_root = root / "profit-output"

    def tearDown(self) -> None:
        self.service.close()
        self.temp.cleanup()

    def test_settings_product_query_import_and_activity_filter(self) -> None:
        settings = self.client.put(
            "/profit-activity/settings",
            json={
                "save_root": str(self.output_root),
                "activity_min_net_profit": 8,
                "activity_profit_rate_threshold": 0.20,
            },
        )
        self.assertEqual(settings.status_code, 200)
        self.assertEqual(Path(settings.json()["save_root"]), self.output_root)

        created = self.client.post(
            "/profit-activity/products",
            json={"site": "US", "skc": "SKC-001", "selling_price": 100, "cost_price": 10, "weight_kg": 0.2},
        )
        self.assertEqual(created.status_code, 200)
        self.assertEqual(created.json()["product"]["skc"], "SKC-001")

        queried = self.client.get("/api/v1/profit-activity/products", params={"site": "US", "skcs": "SKC-001\nUNKNOWN"})
        self.assertEqual(queried.status_code, 200)
        self.assertEqual([row["skc"] for row in queried.json()["products"]], ["SKC-001"])

        import_preview = self.client.post(
            "/profit-activity/products/import/preview",
            files={"file": ("products.xlsx", _workbook_bytes(["SKC", "selling price", "cost", "weight kg"], [["SKC-002", 90, 8, 0.15]]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"site": "US"},
        )
        self.assertEqual(import_preview.status_code, 200)
        preview_data = import_preview.json()
        self.assertEqual(preview_data["summary"]["importable_rows"], 1)
        confirmed = self.client.post(
            "/profit-activity/products/import/confirm",
            json={"import_id": preview_data["import_id"], "selected_row_ids": ["Sheet:2"], "on_conflict": "skip"},
        )
        self.assertEqual(confirmed.status_code, 200)
        self.assertEqual(confirmed.json()["imported"], 1)

        activity_workbook = Workbook()
        price_sheet = activity_workbook.active
        price_sheet.title = "activity_price"
        price_sheet.append(["activity", "SPU ID", "SKC ID", "SKU ID", "activity price"])
        price_sheet.append(["sale", "SPU-1", "SKC-001", "SKU-1", 50])
        price_sheet.append(["sale", "SPU-2", "MISSING", "SKU-2", 50])
        inventory_sheet = activity_workbook.create_sheet("activity_inventory")
        inventory_sheet.append(["activity", "SPU ID", "stock"])
        inventory_sheet.append(["sale", "SPU-1", 10])
        inventory_sheet.append(["sale", "SPU-2", 10])
        activity_stream = BytesIO()
        activity_workbook.save(activity_stream)
        filtered = self.client.post(
            "/profit-activity/activity-filter",
            files={"file": ("activity.xlsx", activity_stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"site": "US", "async_mode": "true"},
        )
        self.assertEqual(filtered.status_code, 200)
        result = filtered.json()
        self.assertEqual(result["status"], "completed")
        self.assertEqual(result["kept_row_count"], 1)
        self.assertEqual(result["removed_row_count"], 1)
        self.assertTrue(Path(result["filtered_path"]).is_file())
        self.assertTrue(Path(result["removed_path"]).is_file())

        download = self.client.get(f"/profit-activity/activity-filter/{result['task_id']}/download", params={"kind": "filtered"})
        self.assertEqual(download.status_code, 200)
        workbook = load_workbook(BytesIO(download.content))
        self.assertEqual(workbook.active.max_row, 2, "the original activity template keeps only the eligible row")
        self.assertEqual(workbook["activity_inventory"].max_row, 2, "orphan inventory rows are removed with their price row")


def _workbook_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


if __name__ == "__main__":
    unittest.main()
