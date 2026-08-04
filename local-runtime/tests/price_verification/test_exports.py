from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from wh_local.price_verification.contracts import (  # noqa: E402
    PluginCommandRequest,
    PriceVerificationActor,
)
from wh_local.price_verification.exports import export_quote_snapshot  # noqa: E402
from wh_local.price_verification.plugin.service import PluginBridgeService  # noqa: E402
from wh_local.price_verification.quote_normalizer import QuoteCounts, QuoteItem, QuotePreview  # noqa: E402
from wh_local.price_verification.quote_service import QuoteService  # noqa: E402
from wh_local.price_verification.repository import (  # noqa: E402
    PriceVerificationRepository,
    QuoteRunRecord,
)


def actor(workspace_id: str) -> PriceVerificationActor:
    return PriceVerificationActor(actor_id=f"user-{workspace_id}", workspace_id=workspace_id)


@pytest.fixture
def repository(tmp_path: Path) -> PriceVerificationRepository:
    repository = PriceVerificationRepository(tmp_path / "runtime.sqlite3")
    yield repository
    repository.close()


@pytest.fixture
def service(repository: PriceVerificationRepository, tmp_path: Path) -> QuoteService:
    return QuoteService(
        repository=repository,
        plugin_bridge=PluginBridgeService(repository=repository),
        output_root=tmp_path / "outputs",
    )


def completed_quote_command(repository: PriceVerificationRepository, skc_id: str):
    session = repository.create_plugin_session(
        workspace_id="A", session_token_hash=(skc_id.encode().hex() * 64)[:64], browser="Edge"
    )
    command = repository.create_command(
        workspace_id="A",
        session_id=session.session_id,
        request=PluginCommandRequest(
            command_type="temu_price_quote_discovery", payload={}, idempotency_key=f"request-{skc_id}"
        ),
    )
    repository.lease_plugin_commands(
        workspace_id="A",
        session_id=session.session_id,
        command_types=("temu_price_quote_discovery",),
        now="2026-08-04T09:00:00+00:00",
        lease_expires_at="2026-08-04T09:02:00+00:00",
        limit=1,
    )
    return repository.record_plugin_result(
        workspace_id="A",
        session_id=session.session_id,
        command_id=command.command_id,
        status="succeeded",
        result={
            "records": [
                {
                    "method": "GET",
                    "url": "https://seller.temu.example/price/quote",
                    "capturedAt": "2026-08-04T09:00:00+00:00",
                    "status": 200,
                    "responseJson": {
                        "data": {
                            "priceReviewItemList": [
                                {
                                    "skcId": skc_id,
                                    "skuId": f"SKU-{skc_id}",
                                    "site": "US",
                                    "supplyPrice": "20.00",
                                    "suggestSupplyPrice": "18.90",
                                    "productTitle": "Test product",
                                    "mainImageUrl": "https://images.example/product.jpg",
                                }
                            ]
                        }
                    },
                }
            ]
        },
        now="2026-08-04T09:00:01+00:00",
    )


def test_export_stays_below_output_root(
    service: QuoteService, repository, tmp_path: Path
) -> None:
    seeded_run_id = service.materialize_completed_command(
        actor("A"), completed_quote_command(repository, "SKC-1")
    ).run_id

    exported = service.export_run(actor("A"), seeded_run_id)

    assert exported.workbook_path.is_relative_to(tmp_path / "outputs")


def test_export_reads_the_persisted_snapshot_and_creates_openable_workbook(
    service: QuoteService, repository
) -> None:
    run = service.materialize_completed_command(
        actor("A"), completed_quote_command(repository, "SKC-1")
    )

    exported = service.export_run(PriceVerificationActor(actor_id="user-A", workspace_id="A"), run.run_id)

    assert exported.workbook_path.name == "normalized_quotes.xlsx"
    assert exported.endpoint_report_path.name == "endpoint_report.md"
    workbook = load_workbook(exported.workbook_path, data_only=True)
    worksheet = workbook.active
    assert worksheet.title == "Normalized Quotes"
    assert worksheet["A2"].value == "SKC-1"
    assert "/price/quote" in exported.endpoint_report_path.read_text(encoding="utf-8")


def test_exporter_rejects_output_paths_that_escape_the_root(tmp_path: Path) -> None:
    run = QuoteRunRecord(
        run_id="../../outside",
        workspace_id="A",
        command_id="command",
        status="succeeded",
        item_count=0,
        adapter_version="",
        captured_at="2026-08-04T09:00:00+00:00",
        created_at="2026-08-04T09:00:00+00:00",
    )

    try:
        export_quote_snapshot(output_root=tmp_path / "outputs", run=run, preview=service_preview())
    except ValueError as error:
        assert "output root" in str(error)
    else:  # pragma: no cover - protects the containment assertion above.
        raise AssertionError("escaped output path was accepted")


def test_export_redacts_and_neutralizes_spreadsheet_formulas(tmp_path: Path) -> None:
    run = QuoteRunRecord(
        run_id="safe-run",
        workspace_id="A",
        command_id="command",
        status="succeeded",
        item_count=1,
        adapter_version="",
        captured_at="2026-08-04T09:00:00+00:00",
        created_at="2026-08-04T09:00:00+00:00",
    )
    preview = QuotePreview(
        quotes=[QuoteItem(skc_id="SKC-1", product_title="=token=never-export")],
        counts=QuoteCounts(1, 0, 1, 0, 0, 0, 0, 0, 0, 0, 1, 1, 0),
        confidence_counts={},
        authenticity_status_counts={},
    )

    exported = export_quote_snapshot(output_root=tmp_path / "outputs", run=run, preview=preview)

    assert load_workbook(exported.workbook_path, data_only=False).active["P2"].value == "'=token=[REDACTED]"


def service_preview():
    return QuotePreview(
        quotes=[],
        counts=QuoteCounts(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
        confidence_counts={},
        authenticity_status_counts={},
    )
