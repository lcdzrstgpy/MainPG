from __future__ import annotations

import csv
import math
import re
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook, load_workbook

from .image_slots import apply_slot_overrides
from .policy import is_safe_external_url


# 店小秘导入默认值（对齐原型 native_product_engine 常量）
DEFAULT_SHIP_DAYS = 2
DECLARED_PRICE_MULTIPLIER = 4
DECLARED_PRICE_MIN_CNY = 150.0
DXM_STOCK_MIN = 0
DXM_STOCK_MAX = 999999
# 店小秘 *重量（g） 导入允许区间上限（校验规则 0.01-99999.9）
DXM_WEIGHT_MAX_G = 99999.9

# 外包装形状/类型（对齐原型 _build_dxm_row：soft → 软包装软物/气泡袋，rigid → 硬包装硬物/纸箱）
_PACKAGE_EXPORT_BY_PROFILE = {
    "rigid_container": ("硬包装硬物", "纸箱"),
}
_PACKAGE_EXPORT_DEFAULT = ("软包装软物", "气泡袋")

# 变种规格轴名称本地映射（对齐原型 §8.2：映射为 Color/Size/Pack/Style/Capacity 等店小秘规格轴）
_VARIANT_AXIS_NAMES = {
    "规格": "Style",
    "规格分类": "Style",
    "款式": "Style",
    "颜色": "Color",
    "颜色分类": "Color",
    "尺寸": "Size",
    "尺码": "Size",
    "型号": "Model",
    "材质": "Material",
    "材料": "Material",
    "套装": "Pack",
    "数量": "Quantity",
    "容量": "Capacity",
    "包装": "Packaging",
    "高度": "Height",
    "长度": "Length",
    "宽度": "Width",
    "形状": "Shape",
}


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "title": ("标题", "商品标题", "商品名称", "产品标题", "title", "name"),
    "skc": ("SKC", "skc", "商品ID", "商品编号"),
    "sku": ("SKU", "sku", "产品货号", "货号"),
    "category": ("类目", "分类", "category"),
    "image_url": ("缩略图链接", "主图 URL", "主图URL", "主图", "图片", "image_url"),
    "source_url": ("链接", "来源", "商品链接", "source_url", "product_link"),
    "price": ("价格", "售价", "最低价格", "建议售价", "price", "price_cny"),
    "description": ("描述", "商品描述", "description"),
    "weight_text": ("*重量（g）", "重量（g）", "*重量(g)", "重量(g)", "重量", "净重"),
    "length_cm": ("*长（cm）", "长（cm）", "*长(cm)", "长(cm)", "长度（cm）", "长"),
    "width_cm": ("*宽（cm）", "宽（cm）", "*宽(cm)", "宽(cm)", "宽度（cm）", "宽"),
    "height_cm": ("*高（cm）", "高（cm）", "*高(cm)", "高(cm)", "高度（cm）", "高"),
}


def _is_http_url(value: Any) -> bool:
    """店小秘图片列只接受无凭据、非本机/内网的公开 HTTP(S) 地址。"""
    return is_safe_external_url(str(value or "").strip())


def _http_urls(values: Any) -> list[str]:
    return [str(value).strip() for value in (values or []) if _is_http_url(value)]


def require_final_public_image_urls(values: list[str]) -> list[str]:
    """Fail closed when a final workbook still contains a local/private image."""
    normalized = [str(value or "").strip() for value in values]
    if any(
        not value.lower().startswith("https://") or not is_safe_external_url(value)
        for value in normalized
    ):
        raise ValueError("final workbook images must be public HTTPS URLs")
    return normalized


def read_product_workbook(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _read_csv(content)
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("only .xlsx, .xlsm or .csv product files are supported")
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [_clean_header(value) for value in next(iterator)]
    except StopIteration as exc:
        raise ValueError("uploaded product workbook is empty") from exc
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(iterator, start=2):
        raw = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if not any(value not in (None, "") for value in raw.values()):
            continue
        rows.append(_normalize_row(raw, row_number))
    workbook.close()
    return rows


# 店小秘导入模板列（与原型程序 native_product_engine.DXM_COLUMNS 一致）
DXM_COLUMNS = [
    "*产品标题",
    "*英文标题",
    "产品描述",
    "产品货号",
    "*变种属性名称一",
    "*变种属性值一",
    "变种属性名称二",
    "变种属性值二",
    "预览图",
    "*申报价格\n(店铺币种)",
    "SKU货号",
    "*长（cm）",
    "*宽（cm）",
    "*高（cm）",
    "*重量（g）",
    "识别码类型",
    "识别码",
    "站外产品链接",
    "*轮播图",
    "*产品素材图",
    "外包装形状",
    "外包装类型",
    "外包装图片",
    "建议售价（USD）",
    "库存",
    "发货时效（天）",
]

DXM_SKU_CLASSIFICATION_COLUMNS = [
    "SKU分类",
    "SKU分类数量",
    "SKU分类单位",
    "独立包装",
    "净含量数值",
    "净含量单位",
    "混合套装类型",
    "SKU分类总数量",
    "SKU分类总数量单位",
    "总净含量",
    "总净含量单位",
    "包装清单",
]


def create_result_workbook(rows: list[dict[str, Any]], destination: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "店小秘导入"
    sheet.append(DXM_COLUMNS + ["*产品分类", "产品分类", "类目路径", "类目ID"] + DXM_SKU_CLASSIFICATION_COLUMNS)
    for row in rows:
        for export_row in _dxm_export_rows(row):
            sheet.append(export_row)
    sheet.freeze_panes = "A2"
    for index, width in enumerate((36, 36, 60, 18, 14, 16, 14, 16, 45, 14, 18, 12, 12, 12, 14, 12, 16, 45, 60, 60, 14, 14, 45, 14, 10, 12), start=1):
        sheet.column_dimensions[_column_letter(index)].width = width
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _dxm_export_rows(row: dict[str, Any]) -> list[list[Any]]:
    """店小秘模板按 SKU 逐行输出：每个来源变种一行，无变种时输出单行。

    店小秘以「变种属性名称+值组合」识别 SKU：来源 1688 数据里同一规格组合可能对应多个
    SKU 货号（价格/库存不同），逐行全量导出会被判定为重复行导致整行拒绝
    （对齐交接文档 §8.4“每个导出规格组合唯一”）。因此按导出的属性组合去重，仅保留首行。
    """
    variant_records = row.get("source_variant_records") or []
    records = [item for item in variant_records if isinstance(item, dict)]
    if not records:
        return [_dxm_single_export_row(row, None)]
    exported: list[list[Any]] = []
    seen: set[tuple[Any, Any, Any, Any]] = set()
    for record in records:
        values = _dxm_single_export_row(row, record)
        # 变种属性名一/值一 + 属性名二/值二（export 行第 4~7 列）
        variant_key = (values[4], values[5], values[6], values[7])
        if variant_key in seen:
            continue
        seen.add(variant_key)
        exported.append(values)
    return exported if exported else [_dxm_single_export_row(row, None)]


def _dxm_single_export_row(row: dict[str, Any], variant: dict[str, Any] | None) -> list[Any]:
    # 预检覆盖（precheck 页保存的标题/描述/图片/核心字段，用户可改可不改，默认保留生成结果）
    preview_overrides = row.get("preview_overrides") or {}
    if not isinstance(preview_overrides, dict):
        preview_overrides = {}
    core_fields = preview_overrides.get("core_fields") or {}
    if not isinstance(core_fields, dict):
        core_fields = {}
    override_carousel = _http_urls(preview_overrides.get("carousel_images"))
    slot_overrides = preview_overrides.get("image_slot_overrides") or {}
    if not isinstance(slot_overrides, dict):
        slot_overrides = {}
    override_main = str(preview_overrides.get("main_image") or "").strip()
    override_detail = _http_urls(preview_overrides.get("detail_images"))

    optimized_title = str(preview_overrides.get("title") or row.get("optimized_title") or "").strip()
    description = str(preview_overrides.get("description") or row.get("description") or "").strip()
    skc = str(row.get("skc") or "").strip()
    sku = str(core_fields.get("sku") or row.get("sku") or skc).strip()
    main_image_url = str(row.get("image_url") or "").strip()
    source_url = str(row.get("source_url") or "").strip()
    source_image_urls = row.get("source_image_urls") or []
    source_detail_image_urls = row.get("source_detail_image_urls") or []
    source_attributes = row.get("source_attributes") or []
    cost = row.get("cost")
    category = str(row.get("category") or "").strip()
    category_path = str(core_fields.get("category_path") or row.get("category_path") or category).strip()
    category_id = str(core_fields.get("category_id") or row.get("category_id") or "").strip()

    # 变种属性值翻译表（来源中文值 → 目标语言显示名，由 service 的 AI 翻译步骤生成）
    value_translations = row.get("variant_value_translations") or {}
    if not isinstance(value_translations, dict):
        value_translations = {}

    # 变种属性：SKU 自己的 attributes 优先，其次取商品级前两条「名称+值」属性
    if variant is not None:
        attributes = variant.get("attributes") or {}
        if not isinstance(attributes, dict):
            attributes = {}
        display_name = str(variant.get("display_name") or "").strip()
        variant_values = []
        for key, value in attributes.items():
            name_text = str(key or "").strip()
            value_text = str(value or "").strip()
            if not name_text or not value_text:
                continue
            # 规格轴名称本地映射 + 属性值翻译（操作员编辑的 display_name 优先）
            export_value = display_name if display_name else value_translations.get(value_text, value_text)
            variant_values.append((_VARIANT_AXIS_NAMES.get(name_text, name_text), export_value))
        variant_sku = str(variant.get("sku_id") or "").strip() or sku
    else:
        variant_values = []
        variant_sku = sku

    if not variant_values:
        # 商品级属性兜底：兼容 list[dict] / dict / list[tuple]，仅取名称+值均完整、非来源类的属性
        attribute_items: list[tuple[Any, Any]] = []
        if isinstance(source_attributes, dict):
            attribute_items = list(source_attributes.items())
        elif isinstance(source_attributes, list):
            for item in source_attributes:
                if isinstance(item, dict):
                    attribute_items.append((item.get("name"), item.get("value")))
                else:
                    try:
                        attribute_items.append((item[0], item[1]))
                    except (TypeError, IndexError, KeyError):
                        continue
        for name, value in attribute_items:
            name_text = str(name or "").strip()
            value_text = str(value or "").strip()
            if not name_text or not value_text or name_text.casefold() in {"来源", "平台", "链接", "图片"}:
                continue
            variant_values.append(
                (_VARIANT_AXIS_NAMES.get(name_text, name_text), value_translations.get(value_text, value_text))
            )
            if len(variant_values) >= 2:
                break

    variant_name_1, variant_value_1, variant_name_2, variant_value_2 = "", "", "", ""
    if variant_values:
        variant_name_1, variant_value_1 = variant_values[0]
        if len(variant_values) > 1:
            variant_name_2, variant_value_2 = variant_values[1]
    if not variant_name_1:
        variant_name_1 = "规格"
    if not variant_value_1:
        # 店小秘 *变种属性值一 必填；无规格值数据时对齐原型 _default_variant_export_value()
        variant_value_1 = "Estándar" if str(row.get("target_language") or "").strip().casefold() == "es" else "Standard"

    # 四宫格落位（对齐交接文档 §11.3）：预览图/素材图=第1张分图；轮播图=4张分图+完整四宫格总览（总览放最后）。
    # 生成图为本地路径（未上传 COS）时店小秘无法访问，仅 http(s) 生成图才可用，否则回退来源 http 图片。
    # 预检覆盖优先：用户改过标题/图片后以覆盖值为准，未改则走原生成/回退逻辑。
    generated_carousel = row.get("carousel_image_paths") or []
    grid_summary_path = str(row.get("grid_image_summary_path") or "").strip()
    detail_image_paths = row.get("detail_image_paths") or []
    generated_images = _http_urls(list(generated_carousel) + [grid_summary_path])
    if slot_overrides:
        # 新版尺寸画布以旧版整组人工轮播为基线再覆盖单槽；总览仍保留在末尾。
        slotted_images = _http_urls(
            [slot.get("value") for slot in apply_slot_overrides(row, preview_overrides)]
        )
        export_images = slotted_images + _http_urls([grid_summary_path])
        carousel = "\n".join(export_images)
        main_image = override_main if _is_http_url(override_main) else next(iter(slotted_images), "")
        material_images = main_image
    elif override_carousel:
        # 纯旧版整组轮播图覆盖保持原语义，避免历史预检数据被悄悄追加图片。
        carousel = "\n".join(override_carousel)
        main_image = override_main if _is_http_url(override_main) else override_carousel[0]
        material_images = main_image
    elif generated_images:
        carousel = "\n".join(generated_images)
        main_image = generated_images[0]
        material_images = generated_images[0]
    else:
        carousel = "\n".join(_http_urls(source_image_urls))
        main_image = main_image_url if _is_http_url(main_image_url) else next(iter(_http_urls(source_image_urls)), "")
        # 店小秘 *产品素材图 为单值列（最大导入1条，对齐原型 DXM_COLUMNS[19]=main_image_url）
        material_images = next(iter(_http_urls(source_detail_image_urls)), "")
        if not material_images:
            material_images = main_image

    # 详情图以 HTML 追加到产品描述（交接文档 §10/§12）；仅追加可外部访问的 http(s) 地址
    # Presence is semantic: an explicit empty array means the operator removed
    # every detail image and must never resurrect generated legacy values.
    detail_sources = (
        override_detail
        if "detail_images" in preview_overrides
        else _http_urls(detail_image_paths)
    )
    detail_html = "".join(f'<img src="{value}" />' for value in detail_sources)
    if detail_html:
        description = f"{description}\n{detail_html}".strip()

    # 物流尺寸/重量与包装（对齐原型 _build_dxm_row：AI 尺寸预估 + 包装形状/类型导出标签）
    dimensions = row.get("product_dimensions") or {}
    if not isinstance(dimensions, dict):
        dimensions = {}
    length = _export_number(core_fields.get("length_cm") if "length_cm" in core_fields else dimensions.get("length_cm"))
    width = _export_number(core_fields.get("width_cm") if "width_cm" in core_fields else dimensions.get("width_cm"))
    height = _export_number(core_fields.get("height_cm") if "height_cm" in core_fields else dimensions.get("height_cm"))
    manual_weight_override = "weight_g" in core_fields
    weight = _export_number(core_fields.get("weight_g") if manual_weight_override else dimensions.get("weight_g"))
    package_shape, package_type = _package_export_values(dimensions)

    # 店小秘体积重校验兜底：变种属性里的尺寸（如 30*20*10cm，店小秘以此算体积重）
    # 与独立长宽高列取体积重（长×宽×高÷6），重量必须大于体积重，否则导入整行被拒。
    source_attr_map = source_attributes if isinstance(source_attributes, dict) else {}
    dimensions_texts: list[Any] = []
    if length not in ("", None) and width not in ("", None) and height not in ("", None):
        dimensions_texts.append(f"{length}*{width}*{height}")
    # 变种属性全部值（不限于导出前两条规格轴）与商品级属性都可能携带尺寸文本
    if variant is not None:
        variant_attributes = variant.get("attributes") or {}
        if isinstance(variant_attributes, dict):
            for attr_value in variant_attributes.values():
                dimensions_texts.append(attr_value)
                # 1688 规格表：变种属性值原文（如【45*50cm】2.5丝，常规款）是规格表 key，
                # 对应 value（如 "25 20 0.50 250 4"）携带完整 长/宽/高/体积/重量。
                spec_row = source_attr_map.get(attr_value)
                if spec_row:
                    dimensions_texts.append(spec_row)
    if isinstance(source_attributes, dict):
        dimensions_texts.extend(source_attributes.values())
    dimensions_texts.extend(value for _, value in variant_values)
    # The precheck value is an operator-confirmed actual weight. Export it
    # exactly as entered instead of silently replacing it with volumetric
    # weight. The import-safety fallback remains for system-generated values.
    if not manual_weight_override:
        weight = _weight_meeting_volumetric(weight, dimensions_texts)

    # 长宽高列兜底：AI 未产出 product_dimensions（长宽高缺失）时，从变种/规格表文本
    # 解析首个三维尺寸填列，保证店小秘 *长/宽/高（cm） 必填列非空。
    if length == "" or width == "" or height == "":
        for text in dimensions_texts:
            parsed_lwh = _parse_dimensions(text)
            if parsed_lwh is not None:
                length = _export_number(parsed_lwh[0])
                width = _export_number(parsed_lwh[1])
                height = _export_number(parsed_lwh[2])
                break

    # 建议售价（对齐原型 _build_dxm_row）：变种建议售价 → 行建议售价 → 来源成本；预检核心字段覆盖优先
    suggested_price = variant.get("suggested_price") if variant else None
    if suggested_price in (None, ""):
        suggested_price = row.get("suggested_price")
    if suggested_price in (None, ""):
        suggested_price = cost
    if core_fields.get("suggested_price") not in (None, ""):
        suggested_price = core_fields.get("suggested_price")

    # 申报价格（对齐原型 _declared_price_for）：
    # 显式申报价 → max(价, 150)；否则 建议售价×4，下限 150；无任何价据 → 150
    declared_price_value = variant.get("declared_price") if variant else None
    if declared_price_value in (None, ""):
        declared_price_value = row.get("declared_price")
    if core_fields.get("declared_price") not in (None, ""):
        declared_price_value = core_fields.get("declared_price")
    parsed_declared = _parse_money(declared_price_value)
    if parsed_declared is not None:
        declared_price_value = max(parsed_declared, DECLARED_PRICE_MIN_CNY)
    else:
        declared_price_value = _declared_price_fallback(suggested_price)
        if declared_price_value is None:
            declared_price_value = DECLARED_PRICE_MIN_CNY

    stock = _normalize_stock(variant.get("stock") if variant else None)
    if stock <= 0:
        stock = _normalize_stock(row.get("stock"))
    if core_fields.get("stock") not in (None, ""):
        stock = _normalize_stock(core_fields.get("stock"))

    return [
        optimized_title,
        optimized_title,
        description,
        skc,
        variant_name_1,
        variant_value_1,
        variant_name_2,
        variant_value_2,
        main_image,
        declared_price_value if declared_price_value not in (None, "") else "",
        variant_sku,
        length,
        width,
        height,
        weight,
        "",  # 识别码类型
        "",  # 识别码
        source_url,
        carousel,
        material_images,
        package_shape,
        package_type,
        "",  # 外包装图片
        suggested_price if suggested_price not in (None, "") else "",
        stock,
        DEFAULT_SHIP_DAYS,  # 发货时效（天）
        category_path,  # *产品分类
        category_path,  # 产品分类
        category_path,  # 类目路径
        category_id,  # 类目ID
        "单品",  # SKU分类
        1,  # SKU分类数量
        "件",  # SKU分类单位
        "", "", "", "", "", "", "", "", "",  # 其余 SKU 分类字段占位
    ]


def _package_export_values(dimensions: dict[str, Any]) -> tuple[str, str]:
    profile = str(dimensions.get("package_profile") or "").strip().lower()
    return _PACKAGE_EXPORT_BY_PROFILE.get(profile, _PACKAGE_EXPORT_DEFAULT)


def _export_number(value: Any) -> Any:
    """把数值型字段转成整数/float，空值保持空串（避免写入 0 掩盖缺失）。"""
    if value in (None, ""):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    if number == int(number):
        return int(number)
    return round(number, 2)


# 尺寸文本模式：如 "30*20*10" / "30×20×10cm" / "40.5*30*20 CM"（1688 变种尺寸属性值）
_DIMENSIONS_PATTERN = re.compile(
    r"(?P<l>\d+(?:\.\d+)?)\s*[*×xX]\s*(?P<w>\d+(?:\.\d+)?)\s*[*×xX]\s*(?P<h>\d+(?:\.\d+)?)"
    r"\s*(?P<unit>cm|厘米|mm|毫米)?",
    re.IGNORECASE,
)

# 1688 规格表行格式：长 宽 高 体积 重量（如 "25 20 0.50 250 4"，取前三个数为长宽高，单位 cm）
_SPACED_DIMENSIONS_PATTERN = re.compile(
    r"(?P<l>\d+(?:\.\d+)?)\s+(?P<w>\d+(?:\.\d+)?)\s+(?P<h>\d+(?:\.\d+)?)"
)


def _parse_dimensions(value: Any) -> tuple[float, float, float] | None:
    """从文本提取 (长, 宽, 高) 厘米；无法识别返回 None。

    支持两种格式：
    1. 乘号分隔三维尺寸（30*20*10cm / 30×20×10 CM），单位 mm/毫米 时换算为厘米（÷10）。
    2. 空格分隔的 1688 规格表行（"25 20 0.50 250 4"，前三个数为长/宽/高 cm）。
    1688 变种尺寸属性常以毫米标注（如 34.5cm 商品写 "345*255*55mm"），
    若不换算会把体积重虚大 1000 倍，导致导出重量超出店小秘允许区间。
    """
    if value in (None, ""):
        return None
    match = _DIMENSIONS_PATTERN.search(str(value))
    if match:
        length = float(match.group("l"))
        width = float(match.group("w"))
        height = float(match.group("h"))
        unit = (match.group("unit") or "").lower()
        if unit in {"mm", "毫米"}:
            length, width, height = length / 10.0, width / 10.0, height / 10.0
    else:
        match = _SPACED_DIMENSIONS_PATTERN.search(str(value))
        if not match:
            return None
        length = float(match.group("l"))
        width = float(match.group("w"))
        height = float(match.group("h"))
    if length <= 0 or width <= 0 or height <= 0:
        return None
    return length, width, height


def _volumetric_weight_g(length: float, width: float, height: float) -> float:
    """店小秘体积重（克）：长×宽×高(cm³) ÷ 6。导入校验要求实际重量大于体积重。"""
    return length * width * height / 6.0


def _weight_meeting_volumetric(weight: Any, dimensions_texts: Sequence[Any]) -> Any:
    """按店小秘体积重校验兜底重量。

    对每段尺寸文本解析长宽高并计算体积重（取最大值），若当前重量缺失或小于等于
    体积重，则提升到「体积重之上 1g」，保证店小秘导入不因「材积重量大于实际重量」
    拒绝整行，且不虚高申报重量。无有效尺寸时原样返回。

    店小秘 *重量（g） 只接受 0.01-99999.9：当兜底值超上限时（例如带 mm 单位但
    未识别的异常尺寸文本使体积重虚大），封顶到上限，避免导出行被区间校验整行拒绝。
    """
    volumetric = 0.0
    for text in dimensions_texts:
        parsed = _parse_dimensions(text)
        if parsed is not None:
            volumetric = max(volumetric, _volumetric_weight_g(*parsed))
    if volumetric <= 0:
        return weight
    if isinstance(weight, (int, float)) and weight > math.ceil(volumetric):
        return weight
    # 体积重为小数时（含小数长宽高），店小秘若将体积重四舍五入到整数再与重量比较，
    # 低于 ceil(vol)+1 的重量（含恰好等于舍入值）会被 "材积重量大于实际重量" 拒绝；
    # ceil(vol)+1 保证重量严格大于任何舍入结果（比最低值多 ≤1g，不虚高）。
    return min(math.ceil(volumetric) + 1, DXM_WEIGHT_MAX_G)


def _normalize_stock(value: Any) -> int:
    amount = _parse_money(value)
    if amount is None:
        return DXM_STOCK_MIN
    return max(DXM_STOCK_MIN, min(DXM_STOCK_MAX, int(amount)))


def _declared_price_fallback(cost: Any) -> float | None:
    """对齐原型 _declared_price_for：无显式申报价时 按成本(CNY)×4，下限 150。"""
    amount = _parse_money(cost)
    if amount is None or amount <= 0:
        return None
    return round(max(amount * DECLARED_PRICE_MULTIPLIER, DECLARED_PRICE_MIN_CNY), 2)


def _parse_money(value: Any) -> float | None:
    """对齐原型 _parse_money：从任意文本中提取第一个非负数字（容忍 $、￥、逗号等）。"""
    text = str(value or "").replace(",", "").strip()
    match = re.search(r"([0-9]+(?:\.[0-9]{1,4})?)", text)
    if not match:
        return None
    try:
        number = float(match.group(1))
    except (TypeError, ValueError):
        return None
    return number if number == number and number not in (float("inf"), float("-inf")) else None


def _column_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def create_error_report(rows: list[dict[str, Any]], destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(["item_id", "draft_id", "title", "status", "reason"])
        for row in rows:
            writer.writerow(
                [row.get("item_id"), row.get("product_draft_id"), row.get("title"), row.get("status"), row.get("reason")]
            )


def create_video_manifest(rows: list[dict[str, Any]], destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(["skc", "title", "source_url", "video_status"])
        for row in rows:
            writer.writerow([row.get("skc"), row.get("optimized_title"), row.get("source_url"), "pending"])


def _read_csv(content: bytes) -> list[dict[str, Any]]:
    decoded = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(StringIO(decoded))
    return [_normalize_row(dict(row), index) for index, row in enumerate(reader, start=2) if any(row.values())]


def _normalize_row(raw: dict[str, Any], row_number: int) -> dict[str, Any]:
    normalized: dict[str, Any] = {"raw_row": raw, "source_row": row_number}
    for target, aliases in HEADER_ALIASES.items():
        normalized[target] = next(
            (raw[alias] for alias in aliases if alias in raw and raw[alias] not in (None, "")), ""
        )
    title = str(normalized["title"] or "").strip()
    if not title:
        normalized["import_warning"] = "missing_title"
    normalized["title"] = title
    normalized["product_name"] = title
    normalized["source_ref"] = str(normalized.get("source_url") or f"workbook-row:{row_number}")
    # 店小秘模板的长/宽/高/重量列：归一为下游可确定性解析的物流原始文本。
    weight = str(normalized.get("weight_text") or "").strip()
    if weight:
        normalized["weight_text"] = weight
    length = str(normalized.get("length_cm") or "").strip()
    width = str(normalized.get("width_cm") or "").strip()
    height = str(normalized.get("height_cm") or "").strip()
    if length and width and height:
        normalized["package_info_text"] = f"{length}x{width}x{height}cm"
    return normalized


def _clean_header(value: Any) -> str:
    return str(value or "").replace("\n", " ").strip()
