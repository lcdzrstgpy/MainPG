from __future__ import annotations

import html
import re
import zipfile
from collections import Counter
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Callable


_HEADER_ALIASES = {
    # 商品唯一标识兼容三种常见来源。数据库仍以 skc 字段保存，避免破坏旧数据；
    # 这里只负责从 Excel 中识别实际使用的 SKU / SKC / SPU / 通用商品 ID 列。
    "product_id": {"商品id", "商品 id", "商品编号", "产品id", "产品 id", "product id", "product_id", "item id", "item_id"},
    "sku": {"sku", "sku id", "sku编号", "sku 编号"},
    "skc": {"skc", "skc id", "skc编号", "skc 编号"},
    "spu": {"spu", "spu id", "spu编号", "spu 编号"},
    "selling_price": {
        "售价", "销售价", "售卖价", "selling price", "price",
        # 核价审核表：核价通过即 Temu 平台售价
        "核价通过", "核价通过价", "审核通过价", "平台核价", "平台价格", "temu价格", "temu售价",
    },
    "cost_price": {
        "成本", "成本价", "采购成本", "cost", "cost price",
        # 核价审核表：国内成本（单价+运输损耗+耗材+运输头程，见 _audit_cost_components）
        "国内成本", "采购成本合计", "成本合计", "商品成本",
    },
    "weight_kg": {
        "重量", "重量kg", "重量 kg", "weight", "weight kg",
        # 核价审核表：实重量在第二行子表头，由 _headers 兜底扫描
        "实重量kg", "实重量 kg", "实重kg", "实重", "实际重量", "实际重量kg",
    },
    "note": {"备注", "说明", "note"},
    "source_url": {
        "货源", "货源链接", "来源链接", "来源 链接", "来源url", "来源 url", "采购链接", "source", "source url",
        "1688产品对应链接", "1688产品链接", "1688链接", "1688货源链接", "产品对应链接",
    },
    "product_image": {
        "商品主图", "主图", "商品图", "预览图", "产品图片", "商品对应图", "商品对应图片",
        "product image", "main image", "sku对应图", "sku对应图片",
        "skc对应图", "skc对应图片", "spu对应图", "spu对应图片",
    },
    "source_image": {"货源图", "采购截图", "source image"},
    "activity_price": {"活动申报价", "活动申报价格", "活动报价", "最低活动价", "activity price", "campaign price"},
    "activity_name": {"活动类型(活动主题)", "活动类型", "活动主题", "activity", "activity name"},
    "site": {"站点", "site", "site code"},
}

# 核价审核表的国内成本被拆成第二行子表头；存在这些子列时，成本按子列求和而非只取“国内成本”合并单元格。
_COST_COMPONENT_HEADERS = {"单价", "运输损耗", "耗材", "运输头程"}

# 仅在首行找不到重量列时，回退扫描第二行子表头（例如 WPS 审核表的“实重量KG”）。
_SECOND_ROW_WEIGHT_HEADERS = {"实重量kg", "实重量 kg", "实重kg", "实重", "实际重量", "实际重量kg"}

# 标准审核表的多货源列：1688产品对应链接/2/3、截图1/2/3、成本2/3，按序号两两配对成货源组。
_SOURCE_URL_NUMBERED_RE = re.compile(r"^1688产品对应链接(\d*)$")
_SOURCE_IMAGE_NUMBERED_RE = re.compile(r"^截图(\d+)$")
_LINK_COST_NUMBERED_RE = re.compile(r"^成本(\d+)$")


def parse_product_workbook(workbook_bytes: bytes, site: str, duplicate_keys: set[tuple[str, str]]) -> list[dict[str, Any]]:
    workbook = _load_workbook(workbook_bytes)
    try:
        rows: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            header_map = _headers(worksheet)
            sheet_site = _worksheet_site(worksheet, header_map, site)
            site_index = header_map.get("site")
            cost_components = _cost_component_indices(worksheet) if header_map.get("cost_price") is not None else []
            for row_number, cells in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                row_site = sheet_site
                if isinstance(site_index, int):
                    cell_site = _cell(cells, site_index)
                    if cell_site is not None and str(cell_site).strip():
                        normalized_site = _normalize_site_value(str(cell_site))
                        if normalized_site in _SITE_CODES:
                            row_site = normalized_site
                values = {field: _cell(cells, index) for field, index in header_map.items() if isinstance(index, int)}
                if not any(str(value or "").strip() for value in values.values()):
                    continue
                product_id, product_id_type = _product_identifier(values)
                blockers: list[str] = []
                warnings: list[str] = []
                if not product_id:
                    blockers.append("missing_product_id")
                numeric = {name: _decimal(values.get(name)) for name in ("selling_price", "cost_price", "weight_kg")}
                if cost_components:
                    numeric["cost_price"] = _sum_decimal(_cell(cells, index) for index in cost_components)
                if not product_id and all(numeric[name] is None for name in ("selling_price", "cost_price", "weight_kg")):
                    # 无 SKC 且没有可用的数值列，视为第二行子表头/说明行，不进入导入结果。
                    continue
                for name, value in numeric.items():
                    if value is None or value <= 0:
                        blockers.append(f"invalid_{name}")
                is_duplicate = (row_site, product_id) in duplicate_keys if product_id else False
                if is_duplicate:
                    warnings.append("duplicate_product_id")
                source_groups = _row_source_groups(cells, header_map)
                source_url = str(values.get("source_url") or (source_groups[0]["source_url"] if source_groups else "") or "").strip()
                rows.append({
                    "row_id": f"{worksheet.title}:{row_number}", "worksheet": worksheet.title,
                    "row_number": row_number, "status": "blocked" if blockers else "ready",
                    "warnings": warnings, "blockers": blockers, "site": row_site, "skc": product_id,
                    "product_id": product_id, "product_id_type": product_id_type, "product_id_label": "商品ID",
                    "selling_price": _number(numeric["selling_price"]), "cost_price": _number(numeric["cost_price"]),
                    "weight_kg": _number(numeric["weight_kg"]), "domestic_fee": None,
                    "note": str(values.get("note") or "").strip(),
                    "source_text": source_url,
                    "source_url": source_url,
                    "source_groups": source_groups,
                    "table_profit": None, "table_profit_rate": None,
                    "has_product_image": False, "has_source_image": False,
                    "product_image_path": "", "source_image_path": "", "is_duplicate": is_duplicate,
                })
        return rows
    finally:
        workbook.close()


def extract_product_workbook_images(workbook_bytes: bytes) -> dict[str, dict[str, list[tuple[str, bytes]]]]:
    """Read images anchored to product rows from an uploaded Excel workbook.

    The workbook format used by employees is not rigid.  An image under a
    recognised main/source-image column is classified accordingly; otherwise
    the first image in a row is treated as the main image and following images
    as source images.  Broken image records are ignored rather than blocking
    the remaining rows.

    WPS 生成的审核表把图片以 DISPIMG 公式嵌入单元格，openpyxl 读不到，
    这里额外解析 xlsx 内部结构补齐这些图片，保证导入后主图/货源图可用。
    """
    workbook = _load_workbook(workbook_bytes)
    dispimg = _dispimg_cell_images(workbook_bytes)
    try:
        result: dict[str, dict[str, list[tuple[str, bytes]]]] = {}
        for worksheet in workbook.worksheets:
            headers = _headers(worksheet)
            product_col = headers.get("product_image")
            source_col = headers.get("source_image")
            source_image_cols = {column: group for group, column in (headers.get("source_images") or {}).items()}
            for image_index, image in enumerate(getattr(worksheet, "_images", []), start=1):
                anchor = getattr(image, "anchor", None)
                origin = getattr(anchor, "_from", None)
                if origin is None:
                    continue
                row_number = int(origin.row) + 1
                column = int(origin.col)
                if row_number < 2:
                    continue
                try:
                    content = image._data()
                except Exception:
                    continue
                if not content:
                    continue
                row_id = f"{worksheet.title}:{row_number}"
                group = result.setdefault(row_id, {"product": [], "source": []})
                if product_col is not None and column == product_col:
                    kind = "product"
                elif column in source_image_cols:
                    kind = f"source_{source_image_cols[column]}"
                elif source_col is not None and column == source_col:
                    kind = "source"
                else:
                    kind = "product" if not group["product"] else "source"
                extension = getattr(image, "format", "png") or "png"
                group.setdefault(kind, []).append((f"excel_{row_number}_{image_index}.{str(extension).lower()}", content))
            # 合并 WPS DISPIMG 公式嵌入的图片（openpyxl 读取不到）
            for row_number, row_images in (dispimg.get(worksheet.title) or {}).items():
                for column, files in row_images.items():
                    row_id = f"{worksheet.title}:{row_number}"
                    group = result.setdefault(row_id, {"product": [], "source": []})
                    if product_col is not None and column == product_col:
                        kind = "product"
                    elif column in source_image_cols:
                        kind = f"source_{source_image_cols[column]}"
                    elif source_col is not None and column == source_col:
                        kind = "source"
                    else:
                        kind = "product" if not group["product"] else "source"
                    group.setdefault(kind, []).extend(files)
        return result
    finally:
        workbook.close()


def _dispimg_cell_images(workbook_bytes: bytes) -> dict[str, dict[int, dict[int, list[tuple[str, bytes]]]]]:
    """解析 WPS DISPIMG 公式嵌入的图片。

    返回 {sheet_title: {row_number: {column: [(filename, bytes)]}}}，行/列均与
    openpyxl 一致（行从 1 开始，列为 0-based）。解析失败时返回空字典，不影响
    openpyxl 常规嵌入图片的读取。
    """
    try:
        zf = zipfile.ZipFile(BytesIO(workbook_bytes))
    except Exception:
        return {}
    try:
        shared = _shared_strings(zf)
        id_to_media = _dispimg_id_to_media(zf)
        if not id_to_media:
            return {}
        try:
            workbook_xml = zf.read("xl/workbook.xml").decode("utf-8", errors="replace")
            workbook_rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
        except KeyError:
            return {}
        rid_to_sheet = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="(worksheets/[^"]+\.xml)"', workbook_rels))
        sheet_names = re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', workbook_xml)
        result: dict[str, dict[int, dict[int, list[tuple[str, bytes]]]]] = {}
        for title, rid in sheet_names:
            sheet_path = rid_to_sheet.get(rid)
            if not sheet_path:
                continue
            try:
                xml = zf.read(f"xl/{sheet_path}").decode("utf-8", errors="replace")
            except KeyError:
                continue
            for row_number, row_cells in _sheet_cells(xml, shared).items():
                for column, value in row_cells.items():
                    if not (isinstance(value, tuple) and value and value[0] == "__DISPIMG__"):
                        continue
                    media = id_to_media.get(value[1])
                    if not media:
                        continue
                    try:
                        content = zf.read(media)
                    except KeyError:
                        continue
                    if not content:
                        continue
                    extension = Path(media).suffix or ".png"
                    result.setdefault(title, {}).setdefault(row_number, {}).setdefault(column, []).append(
                        (f"dispimg_{row_number}_{column}{extension}", content)
                    )
        return result
    except Exception:
        return {}
    finally:
        zf.close()


def _shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        xml = zf.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
    except KeyError:
        return []
    result: list[str] = []
    for si in re.findall(r"<si>(.*?)</si>", xml, re.S):
        parts = re.findall(r"<t[^>]*>(.*?)</t>", si, re.S)
        result.append(html.unescape("".join(parts)))
    return result


def _dispimg_id_to_media(zf: zipfile.ZipFile) -> dict[str, str]:
    """DISPIMG 图片 ID -> zip 内 media 路径（xl/...）。"""
    rid_to_target: dict[str, str] = {}
    try:
        rels = zf.read("xl/_rels/cellimages.xml.rels").decode("utf-8", errors="replace")
    except KeyError:
        return {}
    for rid, target in re.findall(r'<Relationship Id="(rId\d+)"[^>]*Target="([^"]+)"', rels):
        rid_to_target[rid] = target

    id_to_media: dict[str, str] = {}
    try:
        xml = zf.read("xl/cellimages.xml").decode("utf-8", errors="replace")
    except KeyError:
        return id_to_media
    for block in re.findall(r"<etc:cellImage>(.*?)</etc:cellImage>", xml, re.S):
        name_m = re.search(r'<xdr:cNvPr[^>]*name="(ID_[A-Za-z0-9]+)"', block)
        embed_m = re.search(r'<a:blip[^>]*r:embed="(rId\d+)"', block)
        if name_m and embed_m:
            target = rid_to_target.get(embed_m.group(1))
            if target:
                id_to_media[name_m.group(1)] = f"xl/{target}"
    return id_to_media


def _sheet_cells(xml: str, shared: list[str]) -> dict[int, dict[int, object]]:
    """解析工作表 XML，返回 {行号(1-based): {列下标(0-based): 值}}。

    DISPIMG 公式单元格的值为 ("__DISPIMG__", 图片ID)，其余为字符串/数字缓存值。
    """
    cells_by_row: dict[int, dict[int, object]] = {}
    for rownum, body in re.findall(r'<row r="(\d+)"[^>]*>(.*?)</row>', xml, re.S):
        row_cells: dict[int, object] = {}
        for cm in re.finditer(r'<c r="([A-Z]+)\d+"([^>]*?)(?:/>|>(.*?)</c>)', body, re.S):
            col_letters, attrs, content = cm.group(1), cm.group(2), cm.group(3) or ""
            if "t=\"s\"" in attrs:
                m = re.search(r"<v>(\d+)</v>", content)
                row_cells[_column_index(col_letters)] = shared[int(m.group(1))] if m and int(m.group(1)) < len(shared) else None
            elif "t=\"str\"" in attrs:
                img = re.search(r'_xlfn\.DISPIMG\(&quot;(ID_[A-Za-z0-9]+)&quot;,1\)', content)
                if img:
                    row_cells[_column_index(col_letters)] = ("__DISPIMG__", img.group(1))
                else:
                    m = re.search(r"<v>(.*?)</v>", content)
                    row_cells[_column_index(col_letters)] = m.group(1) if m else None
            else:
                m = re.search(r"<v>(.*?)</v>", content)
                row_cells[_column_index(col_letters)] = m.group(1) if m else None
        cells_by_row[int(rownum)] = row_cells
    return cells_by_row


def _column_index(letters: str) -> int:
    """A -> 0, B -> 1 ..."""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - ord("A") + 1)
    return n - 1


def _normalize_site_value(value: str) -> str:
    """把报名表里的中文站名（美国站/哥伦比亚站…）归一化为站点代码，未匹配时原样返回。"""
    site = str(value or "").strip().upper()
    if site in {"US", "CO", "EC", "MX", "BR", "CA"}:
        return site
    aliases = {
        "美国站": "US", "美区": "US", "美国": "US",
        "哥伦比亚站": "CO", "哥伦比亚": "CO",
        "厄瓜多尔站": "EC", "厄瓜多尔": "EC",
        "墨西哥站": "MX", "墨西哥": "MX",
        "巴西站": "BR", "巴西": "BR",
        "加拿大站": "CA", "加拿大": "CA",
    }
    if site in aliases:
        return aliases[site]
    for name, code in aliases.items():
        if name.upper() in site or site in name.upper():
            return code
    return site


_SITE_CODES = {"US", "CO", "EC", "MX", "BR", "CA"}


def _worksheet_site(worksheet, header_map: dict[str, int], fallback: str) -> str:
    """自动识别工作表对应的站点，优先级：sheet 标题 → “站点”列首个非空值 → 兜底默认站点。"""
    title_site = _normalize_site_value(worksheet.title)
    if title_site in _SITE_CODES:
        return title_site
    site_index = header_map.get("site")
    if site_index is not None:
        for cells in worksheet.iter_rows(min_row=2, values_only=True):
            value = _cell(cells, site_index)
            if value is not None and str(value).strip():
                column_site = _normalize_site_value(str(value))
                if column_site in _SITE_CODES:
                    return column_site
                break
    return str(fallback or "US").upper()


class FilterPausedError(InterruptedError):
    """活动过滤被用户暂停时抛出。"""


def filter_activity_workbook(
    workbook_bytes: bytes,
    *,
    site: str,
    evaluate: Callable[[list[str], Decimal], dict[str, Any]],
    should_stop: Callable[[], bool] | None = None,
) -> dict[str, Any]:
    """Filter an activity workbook while retaining its sheet layout and data.

    ``evaluate`` receives ``(candidate_ids, activity_price)`` where
    ``candidate_ids`` is the list of all product identifiers (SPU/SKC/SKU) in
    the row, and returns a stable decision payload containing at least ``keep``,
    ``reason_code``, ``net_profit`` and ``profit_rate``.  The function is
    domain-only: it does not write files or know about HTTP/database concerns.
    """
    workbook = _load_workbook(workbook_bytes)
    try:
        worksheet = _activity_price_sheet(workbook)
        headers = _headers(worksheet)
        product_id_indices = _product_id_indices(headers)
        if not product_id_indices:
            raise ValueError("activity workbook is missing a product ID column")
        price_index = headers.get("activity_price")
        activity_index = headers.get("activity_name")
        site_index = headers.get("site")
        spu_index = headers.get("spu")
        groups: dict[tuple[str, str], list[tuple[int, Decimal | None, list[str]]]] = {}
        spu_by_row: dict[int, str] = {}
        removed_rows: list[dict[str, Any]] = []
        site_mismatch_rows: set[int] = set()

        for row_number, cells in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            candidate_ids = _cell_identifiers(cells, product_id_indices)
            if not candidate_ids:
                continue
            row_site = _normalize_site_value(str(_cell(cells, site_index) or "")) if site_index is not None else ""
            if row_site and row_site != site:
                site_mismatch_rows.add(row_number)
                removed_rows.append(_removed_row(worksheet, row_number, "site_mismatch", None))
                continue
            activity_name = str(_cell(cells, activity_index) or "").strip() if activity_index is not None else ""
            price = _decimal(_cell(cells, price_index)) if price_index is not None else None
            primary_id = candidate_ids[0]
            groups.setdefault((primary_id, activity_name), []).append((row_number, price, candidate_ids))
            if spu_index is not None:
                spu_by_row[row_number] = str(_cell(cells, spu_index) or "").strip()

        decisions: list[dict[str, Any]] = []
        kept_rows: set[int] = set()
        removed_row_numbers: set[int] = set(site_mismatch_rows)
        qualification_counts: Counter[str] = Counter()
        kept_skcs: set[str] = set()
        removed_skcs: set[str] = set()
        kept_spus: set[str] = set()
        kept_activity_spus: set[tuple[str, str]] = set()
        kept_activity_keys: set[tuple[str, str]] = set()
        removed_activity_keys: set[tuple[str, str]] = set()
        # 逐条判定：同一 SKC 在不同活动（活动主题）下申报价不同，每条 SKC×活动×申报价 独立评估，
        # 满足条件的行保留、不符合的行剔除，不再按 SKC 去重汇总、也不取组内最低价。
        for (primary_id, activity_name), entries in groups.items():
            if should_stop and should_stop():
                raise FilterPausedError("filter paused")
            for row_number, price, candidate_ids in entries:
                if price is None or price <= 0:
                    decision: dict[str, Any] = {"keep": False, "decision": "excluded", "reason_code": "invalid_activity_price", "net_profit": None, "profit_rate": None}
                    matched_id = primary_id
                else:
                    decision = evaluate(candidate_ids, price)
                    matched_id = str(decision.get("matched_id") or primary_id)
                decision = {**decision, "skc": matched_id, "product_id": matched_id, "candidate_ids": list(candidate_ids), "activity_name": activity_name, "price": float(price) if price is not None else None, "spu": spu_by_row.get(row_number, "")}
                decisions.append(decision)
                reason = str(decision.get("reason_code") or "unknown")
                qualification_counts[reason] += 1
                if decision.get("keep"):
                    kept_rows.add(row_number)
                    kept_skcs.add(matched_id)
                    kept_activity_keys.add((matched_id, activity_name))
                    spu = spu_by_row.get(row_number, "")
                    if spu:
                        kept_spus.add(spu)
                        kept_activity_spus.add((activity_name, spu))
                else:
                    removed_row_numbers.add(row_number)
                    removed_skcs.add(matched_id)
                    removed_activity_keys.add((matched_id, activity_name))
                    removed_rows.append(_removed_row(worksheet, row_number, reason, decision))

        _delete_rows(worksheet, removed_row_numbers)
        # Activity templates commonly contain a second inventory sheet.  Once
        # a price-row is removed, remove its orphan inventory rows too so the
        # generated workbook remains internally consistent.
        for candidate in workbook.worksheets:
            if candidate is worksheet:
                continue
            candidate_headers = _headers(candidate)
            candidate_spu_index = candidate_headers.get("spu")
            if candidate_spu_index is None:
                continue
            candidate_activity_index = candidate_headers.get("activity_name")
            candidate_remove: set[int] = set()
            for row_number, cells in enumerate(candidate.iter_rows(min_row=2, values_only=True), start=2):
                spu = str(_cell(cells, candidate_spu_index) or "").strip()
                if not spu:
                    continue
                activity_name = str(_cell(cells, candidate_activity_index) or "").strip() if candidate_activity_index is not None else ""
                keep = (activity_name, spu) in kept_activity_spus if candidate_activity_index is not None else spu in kept_spus
                if not keep:
                    candidate_remove.add(row_number)
            _delete_rows(candidate, candidate_remove)
        filtered_bytes = workbook_bytes_from(workbook)
        removed_bytes = _removed_workbook_bytes(removed_rows)
        return {
            "filtered_bytes": filtered_bytes,
            "removed_bytes": removed_bytes,
            "kept_skc_count": len(kept_skcs), "removed_skc_count": len(removed_skcs),
            "kept_row_count": len(kept_rows), "removed_row_count": len(removed_rows),
            "kept_activity_count": len(kept_activity_keys),
            "removed_activity_count": len(removed_activity_keys),
            "qualification_counts": dict(qualification_counts), "removed_rows": removed_rows,
            "activity_decisions": decisions,
            "template_site_summary": {
                "total_price_rows": sum(len(entries) for entries in groups.values()) + len(site_mismatch_rows),
                "site_counts": {site: sum(len(entries) for entries in groups.values())},
                "unique_skc_count_by_site": {site: len({key[0] for key in groups})},
            },
        }
    finally:
        workbook.close()


def new_workbook():
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ValueError("openpyxl is required for Excel support; install openpyxl") from exc
    return Workbook()


def workbook_bytes(workbook) -> bytes:
    return workbook_bytes_from(workbook)


def workbook_bytes_from(workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load_workbook(workbook_bytes: bytes):
    if not workbook_bytes:
        raise ValueError("uploaded workbook is empty")
    try:
        from openpyxl import load_workbook
        # data_only=True 读取公式的缓存结果，否则售价/成本等公式列会拿到公式原文而无法解析数值。
        return load_workbook(BytesIO(workbook_bytes), data_only=True)
    except ImportError as exc:
        raise ValueError("openpyxl is required for Excel support; install openpyxl") from exc
    except Exception as exc:
        raise ValueError("invalid Excel workbook") from exc


def _activity_price_sheet(workbook):
    candidates = []
    for worksheet in workbook.worksheets:
        headers = _headers(worksheet)
        if _product_id_indices(headers):
            candidates.append(("activity_price" in headers, worksheet))
    if not candidates:
        raise ValueError("activity workbook is missing a worksheet with a product ID column")
    return next((sheet for has_price, sheet in candidates if has_price), candidates[0][1])


def _headers(worksheet) -> dict[str, Any]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    result: dict[str, Any] = {}
    for index, raw in enumerate(first_row):
        key = _normalize_header(raw)
        for field, aliases in _HEADER_ALIASES.items():
            if key in aliases and field not in result:
                result[field] = index
    numbered = _numbered_header_columns(first_row)
    if numbered:
        result.update(numbered)
    if result.get("weight_kg") is None:
        # WPS 审核表的“实重量KG”位于第二行子表头，兜底扫描一次。
        second_row = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
        for index, raw in enumerate(second_row):
            if _normalize_header(raw) in _SECOND_ROW_WEIGHT_HEADERS:
                result["weight_kg"] = index
                break
    return result


def _product_identifier(values: dict[str, object | None]) -> tuple[str, str]:
    """Return the first supported product identifier found in one Excel row."""
    for field in ("product_id", "sku", "skc", "spu"):
        value = str(values.get(field) or "").strip()
        if value:
            return value, field
    return "", ""


def _product_id_indices(headers: dict[str, Any]) -> list[int]:
    """返回报名表所有可用的商品标识列下标（SPU/SKC/SKU/商品ID）。

    用户上传的报名表可能同时包含 SPU ID / SKC ID / SKU ID，且不确定会填哪一列，
    因此过滤时把所有标识列都读出来，逐一与产品库的商品ID匹配。
    """
    indices: list[int] = []
    for field in ("product_id", "sku", "skc", "spu"):
        index = headers.get(field)
        if isinstance(index, int) and index not in indices:
            indices.append(index)
    return indices


def _cell_identifiers(cells: tuple[object, ...], indices: list[int]) -> list[str]:
    """按列顺序读取一行的所有商品标识，去空、去重后返回候选标识列表。"""
    seen: list[str] = []
    for index in indices:
        value = str(_cell(cells, index) or "").strip()
        if value and value not in seen:
            seen.append(value)
    return seen


def _numbered_header_columns(first_row: tuple[object, ...]) -> dict[str, dict[int, int]]:
    """扫描标准审核表按序号命名的多货源列，返回按货源组序号索引的列下标映射。

    - source_urls:  {0: 列, 1: 列, ...}  “1688产品对应链接 / 1688产品对应链接2 / …”
    - source_images: {0: 列, 1: 列, ...}  “截图1 / 截图2 / …”
    - link_costs:    {1: 列, 2: 列, ...}  “成本2 / 成本3 / …”（与链接2/3 对应）
    """
    source_urls: dict[int, int] = {}
    source_images: dict[int, int] = {}
    link_costs: dict[int, int] = {}
    for index, raw in enumerate(first_row):
        key = _normalize_header(raw)
        url_match = _SOURCE_URL_NUMBERED_RE.match(key)
        if url_match:
            group = int(url_match.group(1) or "0")
            source_urls.setdefault(group, index)
            continue
        image_match = _SOURCE_IMAGE_NUMBERED_RE.match(key)
        if image_match:
            group = int(image_match.group(1)) - 1
            source_images.setdefault(group, index)
            continue
        cost_match = _LINK_COST_NUMBERED_RE.match(key)
        if cost_match:
            group = int(cost_match.group(1)) - 1
            if group >= 1:  # “成本2/成本3…” 属于链接2/3 的货源成本，成本1 即“国内成本”
                link_costs.setdefault(group, index)
    result: dict[str, dict[int, int]] = {}
    if source_urls:
        result["source_urls"] = source_urls
    if source_images:
        result["source_images"] = source_images
    if link_costs:
        result["link_costs"] = link_costs
    return result


def _cost_component_indices(worksheet) -> list[int]:
    """返回第二行子表头中“单价/运输损耗/耗材/运输头程”的列下标。

    核价审核表的“国内成本”是合并单元格，数值分散在这四列中，导入时应求和。
    """
    second_row = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
    return [index for index, raw in enumerate(second_row) if _normalize_header(raw) in _COST_COMPONENT_HEADERS]


def _sum_decimal(values: Any) -> Decimal | None:
    """对生成器中的单元格值求和；全部为空时返回 None。"""
    total = None
    for value in values:
        parsed = _decimal(value)
        if parsed is None:
            continue
        total = parsed if total is None else total + parsed
    return total


def _row_source_groups(cells: tuple[object, ...], header_map: dict[str, Any]) -> list[dict[str, Any]]:
    """按货源组序号聚合“1688产品对应链接N + 成本N”，生成 [{source_url, cost}] 列表。

    只保留“有链接”或“有非零成本”的组：标准审核表的成本2/3 列常整列为空或 0，
    空组会污染货源组索引（前端按组号加载截图），因此直接丢弃。
    """
    url_cols = header_map.get("source_urls") or {}
    cost_cols = header_map.get("link_costs") or {}
    groups: list[dict[str, Any]] = []
    for group_index in sorted(set(url_cols) | set(cost_cols)):
        url = str(_cell(cells, url_cols.get(group_index)) or "").strip() if group_index in url_cols else ""
        cost = _decimal(_cell(cells, cost_cols.get(group_index))) if group_index in cost_cols else None
        if not url and (cost is None or cost == 0):
            continue
        groups.append({"source_url": url, "cost": _number(cost) if cost is not None else None})
    return groups


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    # 全角标点统一转半角，避免“活动类型(活动主题）”这类表头无法匹配别名。
    text = text.replace("（", "(").replace("）", ")").replace("，", ",").replace("：", ":")
    return re.sub(r"\s+", " ", text)


def _cell(cells: tuple[object, ...], index: int | None) -> object | None:
    return None if index is None or index >= len(cells) else cells[index]


def _decimal(value: object) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        result = Decimal(str(value).strip().replace("¥", "").replace(",", ""))
    except (InvalidOperation, ValueError):
        return None
    return result if result.is_finite() else None


def _number(value: Decimal | None) -> float | None:
    return None if value is None else float(value)


def _delete_rows(worksheet, row_numbers: set[int]) -> None:
    for row_number in sorted((number for number in row_numbers if number > 1), reverse=True):
        worksheet.delete_rows(row_number, 1)


def _removed_row(worksheet, row_number: int, reason_code: str, decision: dict[str, Any] | None) -> dict[str, Any]:
    values = [cell.value for cell in worksheet[row_number]]
    return {
        "worksheet": worksheet.title, "row_number": row_number,
        "values": values, "reason_code": reason_code,
        "net_profit": decision.get("net_profit") if decision else None,
        "profit_rate": decision.get("profit_rate") if decision else None,
    }


def _removed_workbook_bytes(rows: list[dict[str, Any]]) -> bytes:
    workbook = new_workbook()
    worksheet = workbook.active
    worksheet.title = "removed_rows"
    worksheet.append(["worksheet", "row_number", "reason_code", "net_profit", "profit_rate", "original_values"])
    for row in rows:
        worksheet.append([
            row["worksheet"], row["row_number"], row["reason_code"], row["net_profit"], row["profit_rate"],
            " | ".join("" if value is None else str(value) for value in row["values"]),
        ])
    return workbook_bytes_from(workbook)
