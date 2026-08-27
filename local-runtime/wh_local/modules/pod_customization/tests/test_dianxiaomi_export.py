from __future__ import annotations

import io
import json
import sqlite3
from pathlib import Path

import pytest
from fastapi import FastAPI, HTTPException
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image
from pydantic import ValidationError

from wh_local.modules.pod_customization.contracts import (
    BatchCreate,
    BusinessFields,
    Calibration,
    ListingFields,
    NormalizedPoint,
    NormalizedRect,
)
from wh_local.modules.pod_customization.export import (
    _is_public_https_url,
    analyze_dianxiaomi_export,
    build_pod_dianxiaomi_export,
)
from wh_local.modules.pod_customization.repository import PodRepositoryError
from wh_local.modules.pod_customization import router as pod_router_module
from wh_local.modules.pod_customization.router import create_router
from wh_local.modules.pod_customization.service import PodCustomizationService
from wh_local.session import Actor
from wh_local.modules.pod_customization.dianxiaomi import DXM_COLUMNS


class NeverCalledRuntime:
    def submit(self, *_args, **_kwargs):
        raise AssertionError("export must not call AI")


class NeverCalledTitleRuntime:
    configured = True

    def submit(self, *_args, **_kwargs):
        raise AssertionError("export must not call title AI")


def _png() -> bytes:
    output = io.BytesIO()
    Image.new("RGB", (320, 240), "#eee9df").save(output, "PNG")
    return output.getvalue()


def _actor(user_id: str = "operator-1") -> Actor:
    return Actor(id=user_id, username=user_id, role="operator", workspace_id="workspace-a")


def _service(tmp_path: Path) -> PodCustomizationService:
    return PodCustomizationService(
        tmp_path / "workbench.sqlite3",
        tmp_path / "assets",
        NeverCalledRuntime(),
        title_runtime=NeverCalledTitleRuntime(),
        start_workers=False,
    )


def _listing(
    *, title_mode: str = "long", skus: list[dict[str, object]] | None = None
) -> ListingFields:
    return ListingFields(
        declared_price=18.5,
        suggested_price_usd=29.99,
        category_name="家居收纳 > 包袋",
        title_mode=title_mode,
        skus=skus
        or [
            {
                "name": "Default SKU",
                "length_cm": 30,
                "width_cm": 20,
                "height_cm": 10,
                "weight_g": 450,
            }
        ],
    )


def test_listing_fields_requires_sku_specific_positive_dimensions_and_weight() -> None:
    fields = ListingFields(
        declared_price=18.5,
        suggested_price_usd=29.99,
        category_name="家居收纳 > 包袋",
        skus=[
            {
                "name": "CT-BLACK",
                "length_cm": 30,
                "width_cm": 20,
                "height_cm": 10,
                "weight_g": 450,
            }
        ],
    )

    assert fields.model_dump()["skus"] == [
        {
            "name": "CT-BLACK",
            "length_cm": 30.0,
            "width_cm": 20.0,
            "height_cm": 10.0,
            "weight_g": 450.0,
        }
    ]
    assert "weight_g" not in fields.model_dump()
    with pytest.raises(ValidationError):
        ListingFields(
            declared_price=18.5,
            suggested_price_usd=29.99,
            category_name="家居收纳 > 包袋",
            skus=[],
        )
    with pytest.raises(ValidationError):
        ListingFields(
            declared_price=18.5,
            suggested_price_usd=29.99,
            category_name="家居收纳 > 包袋",
            skus=[
                {
                    "name": " ",
                    "length_cm": 30,
                    "width_cm": 20,
                    "height_cm": 10,
                    "weight_g": 0,
                }
            ],
        )


def test_export_expands_new_skus_with_their_own_dimensions_and_last_scene_preview() -> None:
    image_urls = {
        "hero": "https://images.example.com/pod/1/hero.png",
        "detail_a": "https://images.example.com/pod/1/detail-a.png",
        "detail_b": "https://images.example.com/pod/1/detail-b.png",
        "lifestyle": "https://images.example.com/pod/1/final-scene.png",
    }
    batch = {
        "batch_id": "new-sku-batch",
        "requested_count": 1,
        "status": "completed",
        "business_fields": {"product_category": "Home > Bags"},
        "listing_fields": {
            "declared_price": 18.5,
            "suggested_price_usd": 29.99,
            "category_name": "家居收纳 > 包袋",
            "skus": [
                {
                    "name": "CT-BLACK",
                    "length_cm": 30,
                    "width_cm": 20,
                    "height_cm": 10,
                    "weight_g": 450,
                },
                {
                    "name": "CT-SAND",
                    "length_cm": 40,
                    "width_cm": 25,
                    "height_cm": 15,
                    "weight_g": 650,
                },
            ],
        },
        "items": [
            {"style_index": 1, "role": role, "status": "completed", "public_url": url}
            for role, url in image_urls.items()
        ],
    }

    exported = build_pod_dianxiaomi_export(
        batch,
        {1: {"title": "Coastal Tote", "english_title": "Coastal Tote", "description": "Description"}},
    )
    workbook = load_workbook(io.BytesIO(exported.content), data_only=True)
    try:
        rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    assert [row[3] for row in rows] == ["POD-001", "POD-001"]
    assert [row[4] for row in rows] == ["尺寸", "尺寸"]
    assert [row[5] for row in rows] == ["CT-BLACK", "CT-SAND"]
    assert [row[10] for row in rows] == ["CT-BLACK", "CT-SAND"]
    assert [row[11:14] for row in rows] == [(30, 20, 10), (40, 25, 15)]
    assert [row[14] for row in rows] == [450, 650]
    assert [row[8] for row in rows] == [image_urls["lifestyle"], image_urls["lifestyle"]]


def test_export_never_uses_a_chinese_sku_name_as_a_product_or_sku_code() -> None:
    image_urls = {
        "hero": "https://images.example.com/pod/1/hero.png",
        "detail_a": "https://images.example.com/pod/1/detail-a.png",
        "detail_b": "https://images.example.com/pod/1/detail-b.png",
        "lifestyle": "https://images.example.com/pod/1/final-scene.png",
    }
    batch = {
        "batch_id": "chinese-sku-code-batch",
        "requested_count": 1,
        "status": "completed",
        "business_fields": {"product_category": "Home > Bags"},
        "listing_fields": {
            "declared_price": 18.5,
            "suggested_price_usd": 29.99,
            "category_name": "家居收纳 > 包袋",
            "skus": [{"name": "标题款", "length_cm": 30, "width_cm": 20, "height_cm": 10, "weight_g": 450}],
        },
        "items": [
            {"style_index": 1, "role": role, "status": "completed", "public_url": url}
            for role, url in image_urls.items()
        ],
    }

    exported = build_pod_dianxiaomi_export(
        batch,
        {1: {"title": "Coastal Tote", "english_title": "Coastal Tote", "description": "Description"}},
    )
    workbook = load_workbook(io.BytesIO(exported.content), data_only=True)
    try:
        row = next(workbook.active.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    assert row[3] == "POD-001"
    assert row[4] == "尺寸"
    assert row[5] == "标题款"
    assert row[10] == "SKU-001-01"


def test_export_legacy_snapshot_without_skus_uses_global_dimensions_and_sku_names() -> None:
    image_urls = {
        "hero": "https://images.example.com/pod/1/hero.png",
        "detail_a": "https://images.example.com/pod/1/detail-a.png",
        "detail_b": "https://images.example.com/pod/1/detail-b.png",
        "lifestyle": "https://images.example.com/pod/1/final-scene.png",
    }
    batch = {
        "batch_id": "legacy-sku-batch",
        "requested_count": 1,
        "status": "completed",
        "business_fields": {"product_category": "Home > Bags"},
        "listing_fields": {
            "declared_price": 18.5,
            "suggested_price_usd": 29.99,
            "length_cm": 30,
            "width_cm": 20,
            "height_cm": 10,
            "weight_g": 450,
            "category_name": "家居收纳 > 包袋",
            "sku_names": ["CT-BLACK", "CT-SAND"],
        },
        "items": [
            {"style_index": 1, "role": role, "status": "completed", "public_url": url}
            for role, url in image_urls.items()
        ],
    }

    exported = build_pod_dianxiaomi_export(
        batch,
        {1: {"title": "Coastal Tote", "english_title": "Coastal Tote", "description": "Description"}},
    )
    workbook = load_workbook(io.BytesIO(exported.content), data_only=True)
    try:
        rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    assert [row[3] for row in rows] == ["POD-001", "POD-001"]
    assert [row[4] for row in rows] == ["尺寸", "尺寸"]
    assert [row[5] for row in rows] == ["CT-BLACK", "CT-SAND"]
    assert [row[10] for row in rows] == ["CT-BLACK", "CT-SAND"]
    assert [row[11:14] for row in rows] == [(30, 20, 10), (30, 20, 10)]


def _batch(
    service: PodCustomizationService,
    actor: Actor,
    *,
    count: int = 2,
    title_mode: str = "long",
    skus: list[dict[str, object]] | None = None,
) -> dict:
    template = service.upload_template(actor, name="Scene", filename="scene.png", content=_png())
    service.update_template_calibration(
        actor,
        template["id"],
        Calibration(
            mask=NormalizedRect(x=0.2, y=0.2, width=0.6, height=0.6),
            anchor=NormalizedPoint(x=0.5, y=0.5),
        ),
    )
    return service.create_batch(
        actor,
        BatchCreate(
            template_id=template["id"],
            count=count,
            business_fields=BusinessFields(product_name="Canvas tote", product_category="Home > Bags"),
            listing_fields=_listing(title_mode=title_mode, skus=skus),
        ),
        enqueue=False,
    )


def _complete_style(
    service: PodCustomizationService,
    batch_id: str,
    style_index: int,
    *,
    urls: tuple[str, str, str, str] | None = None,
    roles: tuple[str, str, str, str] = ("hero", "detail_a", "detail_b", "lifestyle"),
) -> None:
    urls = urls or tuple(
        f"https://images.example.com/pod/{style_index}/{role}.png" for role in roles
    )
    with sqlite3.connect(service.database_path) as connection:
        rows = connection.execute(
            """SELECT result_id, variant_index FROM pod_customization_style_grid_results
               WHERE batch_id = ? AND style_index = ? ORDER BY variant_index""",
            (batch_id, style_index),
        ).fetchall()
        for (result_id, _variant_index), role, url in zip(rows, roles, urls, strict=True):
            connection.execute(
                """UPDATE pod_customization_style_grid_results
                   SET status = 'completed', pattern_asset_id = 'pattern', composite_asset_id = 'composite'
                   WHERE result_id = ?""",
                (result_id,),
            )
            connection.execute(
                """INSERT INTO pod_customization_style_grid_publications
                   (result_id, role, public_url, updated_at) VALUES (?, ?, ?, 'now')""",
                (result_id, role, url),
            )


def _settle(service: PodCustomizationService, batch_id: str, status: str = "completed") -> None:
    with sqlite3.connect(service.database_path) as connection:
        connection.execute("UPDATE pod_customization_batches SET status = ? WHERE batch_id = ?", (status, batch_id))


def test_service_exports_exact_42_cell_row_and_skips_invalid_styles(tmp_path: Path) -> None:
    service = _service(tmp_path)
    actor = _actor()
    batch = _batch(service, actor)
    _complete_style(service, batch["id"], 1)
    _complete_style(
        service,
        batch["id"],
        2,
        urls=(
            "http://images.example.com/hero.png",
            "https://images.example.com/a.png",
            "https://images.example.com/b.png",
            "https://images.example.com/life.png",
        ),
    )
    _settle(service, batch["id"], "partial_failure")
    service.repository.upsert_style_copy(
        batch["id"], actor.workspace_id, actor.id, 1,
        title="Coastal Tote", english_title="Coastal Canvas Tote", description="Carry calm everywhere.",
    )

    exported = service.export_dianxiaomi(actor, batch["id"])

    assert exported.exported_style_count == 1
    assert exported.skipped_style_count == 1
    assert exported.filename == f"pod_dxm_{batch['id'][:8]}.xlsx"
    workbook = load_workbook(io.BytesIO(exported.content), data_only=True)
    try:
        values = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    assert list(values[0]) == DXM_COLUMNS
    row = list(values[1])
    assert len(row) == 42
    assert row[:6] == [
        "Coastal Tote", "Coastal Tote",
        'Carry calm everywhere.\n<img src="https://images.example.com/pod/1/lifestyle.png" />\n<img src="https://images.example.com/pod/1/detail_a.png" />\n<img src="https://images.example.com/pod/1/detail_b.png" />\n<img src="https://images.example.com/pod/1/hero.png" />',
        "POD-001", "尺寸", "Default SKU",
    ]
    assert row[10] == "SKU-001-01"
    assert row[6:9] == [None, None, "https://images.example.com/pod/1/lifestyle.png"]
    assert row[9:15] == [18.5, "SKU-001-01", 30, 20, 10, 450]
    assert row[15:18] == [None, None, None]
    assert row[18] == "\n".join(
        [
            "https://images.example.com/pod/1/lifestyle.png",
            "https://images.example.com/pod/1/detail_a.png",
            "https://images.example.com/pod/1/detail_b.png",
            "https://images.example.com/pod/1/hero.png",
        ]
    )
    assert row[19:24] == ["https://images.example.com/pod/1/hero.png", None, None, None, 29.99]
    assert row[24:30] == [None, None, "家居收纳 > 包袋", "家居收纳 > 包袋", "家居收纳 > 包袋", None]
    assert row[30:33] == ["单品", 1, "件"]
    assert row[33:] == [None] * 9


def test_service_export_repeats_each_style_for_saved_skus_and_uses_last_scene_as_preview(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path)
    actor = _actor()
    batch = _batch(
        service,
        actor,
        count=1,
        skus=[
            {"name": "CT-BLACK", "length_cm": 30, "width_cm": 20, "height_cm": 10, "weight_g": 450},
            {"name": "CT-SAND", "length_cm": 40, "width_cm": 25, "height_cm": 15, "weight_g": 650},
        ],
    )
    _complete_style(
        service,
        batch["id"],
        1,
        urls=(
            "https://images.example.com/pod/1/hero.png",
            "https://images.example.com/pod/1/detail-a.png",
            "https://images.example.com/pod/1/detail-b.png",
            "https://images.example.com/pod/1/final-scene.png",
        ),
    )
    _settle(service, batch["id"])
    service.repository.upsert_style_copy(
        batch["id"], actor.workspace_id, actor.id, 1,
        title="Coastal Tote", english_title="Coastal Canvas Tote", description="Carry calm everywhere.",
    )
    exported = service.export_dianxiaomi(actor, batch["id"])
    workbook = load_workbook(io.BytesIO(exported.content), data_only=True)
    try:
        rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    assert exported.exported_style_count == 1
    assert service.get_batch(actor, batch["id"])["listing_fields"]["skus"] == [
        {"name": "CT-BLACK", "length_cm": 30.0, "width_cm": 20.0, "height_cm": 10.0, "weight_g": 450.0},
        {"name": "CT-SAND", "length_cm": 40.0, "width_cm": 25.0, "height_cm": 15.0, "weight_g": 650.0},
    ]
    assert [row[3] for row in rows] == ["POD-001", "POD-001"]
    assert [row[4] for row in rows] == ["尺寸", "尺寸"]
    assert [row[5] for row in rows] == ["CT-BLACK", "CT-SAND"]
    assert [row[10] for row in rows] == ["CT-BLACK", "CT-SAND"]
    assert [row[11:14] for row in rows] == [(30, 20, 10), (40, 25, 15)]
    assert [row[14] for row in rows] == [450, 650]
    assert [row[8] for row in rows] == [
        "https://images.example.com/pod/1/final-scene.png",
        "https://images.example.com/pod/1/final-scene.png",
    ]


def test_service_export_uses_selected_short_title_for_both_title_columns(tmp_path: Path) -> None:
    service = _service(tmp_path)
    actor = _actor()
    batch = _batch(service, actor, count=1, title_mode="short")
    _complete_style(service, batch["id"], 1)
    _settle(service, batch["id"])
    service.repository.upsert_style_copy(
        batch["id"], actor.workspace_id, actor.id, 1,
        title="Long descriptive product title", english_title="Short product title", description="Description",
    )

    workbook = load_workbook(io.BytesIO(service.export_dianxiaomi(actor, batch["id"]).content), data_only=True)
    try:
        row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    finally:
        workbook.close()

    assert row[0] == row[1] == "Short product title"


def test_historical_listing_snapshot_without_title_mode_defaults_to_long_title(tmp_path: Path) -> None:
    service = _service(tmp_path)
    actor = _actor()
    batch = _batch(service, actor, count=1)
    _complete_style(service, batch["id"], 1)
    _settle(service, batch["id"])
    service.repository.upsert_style_copy(
        batch["id"], actor.workspace_id, actor.id, 1,
        title="Historical long title", english_title="Historical short title", description="Description",
    )
    with sqlite3.connect(service.database_path) as connection:
        raw = connection.execute(
            "SELECT listing_fields_json FROM pod_customization_batches WHERE batch_id = ?",
            (batch["id"],),
        ).fetchone()[0]
        listing_fields = json.loads(raw)
        listing_fields.pop("title_mode")
        connection.execute(
            "UPDATE pod_customization_batches SET listing_fields_json = ? WHERE batch_id = ?",
            (json.dumps(listing_fields), batch["id"]),
        )

    workbook = load_workbook(io.BytesIO(service.export_dianxiaomi(actor, batch["id"]).content), data_only=True)
    try:
        row = list(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    finally:
        workbook.close()

    assert row[0] == row[1] == "Historical long title"


@pytest.mark.parametrize(
    ("status", "copy", "expected"),
    [("queued", True, "active"), ("completed", False, "copy")],
)
def test_service_export_blocks_active_or_missing_copy(
    tmp_path: Path, status: str, copy: bool, expected: str
) -> None:
    service = _service(tmp_path)
    actor = _actor()
    batch = _batch(service, actor, count=1)
    _complete_style(service, batch["id"], 1)
    _settle(service, batch["id"], status)
    if copy:
        service.repository.upsert_style_copy(
            batch["id"], actor.workspace_id, actor.id, 1,
            title="Title", english_title="English", description="Description",
        )

    with pytest.raises(PodRepositoryError, match=expected) as raised:
        service.export_dianxiaomi(actor, batch["id"])
    assert raised.value.status_code == 409


def test_completed_style_without_copy_is_skipped_when_another_style_is_exportable(tmp_path: Path) -> None:
    service = _service(tmp_path)
    actor = _actor()
    batch = _batch(service, actor, count=2)
    _complete_style(service, batch["id"], 1)
    _complete_style(service, batch["id"], 2)
    _settle(service, batch["id"])
    service.repository.upsert_style_copy(
        batch["id"], actor.workspace_id, actor.id, 1,
        title="Title", english_title="English", description="Description",
    )

    payload = service.get_batch(actor, batch["id"])
    exported = service.export_dianxiaomi(actor, batch["id"])

    assert payload["dianxiaomi_export"] == {
        "ready": True,
        "exportable_style_count": 1,
        "skipped_style_count": 1,
        "block_reason": None,
    }
    assert exported.exported_style_count == 1
    assert exported.skipped_style_count == 1


@pytest.mark.parametrize(
    "malformed_copy",
    [
        None,
        {},
        {"title": "Title", "english_title": "English"},
        {"title": " ", "english_title": "English", "description": "Description"},
        {"title": "Title", "english_title": 123, "description": "Description"},
    ],
)
def test_malformed_or_blank_style_copy_blocks_as_missing_without_key_error(
    tmp_path: Path, malformed_copy: object
) -> None:
    service = _service(tmp_path)
    actor = _actor()
    batch = _batch(service, actor, count=1)
    _complete_style(service, batch["id"], 1)
    _settle(service, batch["id"])
    stored_batch = service.repository.get_batch(batch["id"], actor.workspace_id, actor.id)
    copies = {1: malformed_copy}

    analysis = analyze_dianxiaomi_export(stored_batch, copies)  # type: ignore[arg-type]

    assert analysis.exportable_styles == {}
    assert analysis.skipped_style_count == 1
    assert analysis.block_reason == "style_copy_missing"
    with pytest.raises(ValueError, match="style_copy_missing"):
        build_pod_dianxiaomi_export(stored_batch, copies)  # type: ignore[arg-type]


def test_malformed_style_copy_is_skipped_when_another_style_has_valid_copy(tmp_path: Path) -> None:
    service = _service(tmp_path)
    actor = _actor()
    batch = _batch(service, actor, count=2)
    _complete_style(service, batch["id"], 1)
    _complete_style(service, batch["id"], 2)
    _settle(service, batch["id"])
    stored_batch = service.repository.get_batch(batch["id"], actor.workspace_id, actor.id)
    copies = {
        1: {"title": "Title", "english_title": "English", "description": "\t"},
        2: {"title": "Valid Title", "english_title": "Valid English", "description": "Valid copy"},
    }

    analysis = analyze_dianxiaomi_export(stored_batch, copies)
    exported = build_pod_dianxiaomi_export(stored_batch, copies)

    assert set(analysis.exportable_styles) == {2}
    assert analysis.skipped_style_count == 1
    assert analysis.block_reason is None
    assert exported.exported_style_count == 1
    assert exported.skipped_style_count == 1
    workbook = load_workbook(io.BytesIO(exported.content), data_only=True)
    try:
        assert workbook.active.cell(2, 1).value == "Valid Title"
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "url",
    [
        "https://2130706433/image.png",
        "https://127.1/image.png",
        "https://0x7f000001/image.png",
        "https://localhost./image.png",
        'https://images.example.com/image\" onerror=\"alert(1).png',
        "https://images.example.com/image' onerror='alert(1).png",
        "https://images.example.com/<script>.png",
        "https://images.example.com/image name.png",
        "https://images.example.com/image%22%20onerror%3Dalert(1).png",
        "https://images.example.com/image%27.png",
        "https://images.example.com/image%3Cscript%3E.png",
        "https://images.example.com/image%09name.png",
        "https://images.example.com/image%0Aname.png",
        "https://１２７。０。０。１/image.png",
        "https://127。0。0。1/image.png",
        "https://intranet/image.png",
        "https://assets.corp.internal/image.png",
        "https://printer.office.lan/image.png",
        "https://printer.localdomain/image.png",
        "https://router.home.arpa/image.png",
        "https://assets.invalid/image.png",
        "https://assets.example.test/image.png",
        "https://images.example.com\\@127.0.0.1/image.png",
        "https://images.example.com/path\\image.png",
        "https://127.0.0.1.nip.io/image.png",
        "https://127-0-0-1.sslip.io/image.png",
    ],
)
def test_public_https_url_rejects_legacy_private_hosts_and_unsafe_attribute_text(url: str) -> None:
    assert not _is_public_https_url(url)


def test_public_https_url_preserves_normal_cos_urls() -> None:
    assert _is_public_https_url(
        "https://pod-bucket-1250000000.cos.ap-guangzhou.myqcloud.com/pod/hero-1.png?version=2"
    )


def test_public_https_url_accepts_canonical_public_idna_hosts() -> None:
    assert _is_public_https_url("https://例子.公司.cn/pod/hero.png")


def test_batch_payload_reports_export_readiness_and_zero_exportable_block(tmp_path: Path) -> None:
    service = _service(tmp_path)
    actor = _actor()
    batch = _batch(service, actor, count=1)
    _settle(service, batch["id"])
    service.repository.upsert_style_copy(
        batch["id"], actor.workspace_id, actor.id, 1,
        title="Title", english_title="English", description="Description",
    )

    payload = service.get_batch(actor, batch["id"])

    assert payload["listing_fields"] == _listing().model_dump()
    assert payload["dianxiaomi_export"] == {
        "ready": False,
        "exportable_style_count": 0,
        "skipped_style_count": 1,
        "block_reason": "no_exportable_styles",
    }
    with pytest.raises(PodRepositoryError, match="exportable") as raised:
        service.export_dianxiaomi(actor, batch["id"])
    assert raised.value.status_code == 409


def test_export_endpoint_returns_xlsx_headers_and_preserves_404_ownership(tmp_path: Path) -> None:
    app = FastAPI()
    router = create_router(
        tmp_path / "workbench.sqlite3",
        tmp_path / "assets",
        NeverCalledRuntime(),
        title_runtime=NeverCalledTitleRuntime(),
        start_workers=False,
    )
    app.include_router(router)
    client = TestClient(app)
    service = getattr(router, "pod_customization_service")
    actor = Actor(id="local-demo-admin", username="local-demo", role="admin", workspace_id="default")
    batch = _batch(service, actor, count=1)
    _complete_style(service, batch["id"], 1)
    _settle(service, batch["id"])
    service.repository.upsert_style_copy(
        batch["id"], actor.workspace_id, actor.id, 1,
        title="Title", english_title="English", description="Description",
    )

    response = client.get(
        f"/api/pod-customization/batches/{batch['id']}/exports/dianxiaomi",
        headers={"Authorization": "Bearer dev-admin-token"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        f'attachment; filename="pod_dxm_{batch["id"][:8]}.xlsx"'
    )
    assert response.headers["x-pod-exported-styles"] == "1"
    assert response.headers["x-pod-skipped-styles"] == "0"
    export_id = response.headers["x-pod-export-id"]
    assert len(export_id) == 32
    assert response.content.startswith(b"PK")

    history = client.get(
        f"/api/pod-customization/batches/{batch['id']}/exports",
        headers={"Authorization": "Bearer dev-admin-token"},
    )
    assert history.status_code == 200
    assert history.json() == {
        "exports": [
            {
                "id": export_id,
                "batch_id": batch["id"],
                "file_name": f"pod_dxm_{batch['id'][:8]}.xlsx",
                "format": "dianxiaomi_xlsx",
                "exported_count": 1,
                "skipped_count": 0,
                "created_at": history.json()["exports"][0]["created_at"],
            }
        ],
        "total": 1,
    }

    missing = client.get(
        "/api/pod-customization/batches/not-owned/exports/dianxiaomi",
        headers={"Authorization": "Bearer dev-admin-token"},
    )
    assert missing.status_code == 404


def test_export_history_is_owner_private_and_contains_no_payload_or_credentials(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path)
    owner = _actor("owner-1")
    batch = _batch(service, owner, count=1)
    _complete_style(service, batch["id"], 1)
    _settle(service, batch["id"])
    service.repository.upsert_style_copy(
        batch["id"], owner.workspace_id, owner.id, 1,
        title="Title", english_title="English", description="Description",
    )

    first = service.export_dianxiaomi(owner, batch["id"])
    second = service.export_dianxiaomi(owner, batch["id"])
    history = service.list_exports(owner, batch["id"])

    assert first.export_id != second.export_id
    assert [record["id"] for record in history["exports"]] == [
        second.export_id,
        first.export_id,
    ]
    assert history["total"] == 2
    serialized = json.dumps(history)
    assert "PK" not in serialized
    assert "token" not in serialized.lower()
    assert "key" not in serialized.lower()
    with pytest.raises(PodRepositoryError) as wrong_owner:
        service.list_exports(_actor("owner-2"), batch["id"])
    assert wrong_owner.value.status_code == 404
    with pytest.raises(PodRepositoryError) as wrong_workspace:
        service.list_exports(
            Actor(id=owner.id, username=owner.username, role="operator", workspace_id="workspace-b"),
            batch["id"],
        )
    assert wrong_workspace.value.status_code == 404


def test_export_endpoint_requires_the_dedicated_export_permission(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    app = FastAPI()
    router = create_router(
        tmp_path / "workbench.sqlite3",
        tmp_path / "assets",
        NeverCalledRuntime(),
        title_runtime=NeverCalledTitleRuntime(),
        start_workers=False,
    )
    app.include_router(router)

    def deny_export(_actor: Actor, permission: str, _database_path: Path) -> None:
        assert permission == "pod_customization.export"
        raise HTTPException(status_code=403, detail="permission denied")

    monkeypatch.setattr(pod_router_module, "require_permission", deny_export)
    response = TestClient(app).get(
        "/api/pod-customization/batches/not-visible/exports/dianxiaomi",
        headers={"Authorization": "Bearer dev-admin-token"},
    )

    assert response.status_code == 403


def test_short_title_export_rejects_noncompliant_copy_instead_of_using_it() -> None:
    from wh_local.modules.pod_customization.export import _build_row

    with pytest.raises(ValueError, match="english_title contains prohibited term"):
        _build_row(
            1,
            {
                "hero": "https://images.example.com/hero.png",
                "detail_a": "https://images.example.com/a.png",
                "detail_b": "https://images.example.com/b.png",
                "lifestyle": "https://images.example.com/lifestyle.png",
            },
            {
                "title": (
                    "Coastal Botanical Canvas Tote with Ocean Fern Artwork and Layered Ink Details "
                    "for Everyday Home Office Studio Carry"
                ),
                "english_title": "Temu exclusive canvas tote",
                "description": "A factual canvas tote description for everyday use.",
            },
            {"product_category": "tote bag"},
            {
                "title_mode": "short",
                "declared_price": 20,
                "suggested_price_usd": 30,
                "length_cm": 20,
                "width_cm": 10,
                "height_cm": 5,
                "weight_g": 300,
                "category_name": "家居收纳 > 包袋",
            },
        )
