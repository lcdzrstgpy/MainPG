from __future__ import annotations

import csv
import io
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree


MAX_DOCUMENT_BYTES = 10 * 1024 * 1024
MAX_EXTRACTED_CHARS = 24_000
MAX_SPREADSHEET_ROWS = 5_000
MAX_SPREADSHEET_COLUMNS = 30
MAX_SHEETS = 20

DOCUMENT_MIME_TYPES = {
    ".txt": "text/plain",
    ".csv": "text/csv",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}


def extract_local_document(filename: str, content: bytes) -> tuple[str, str]:
    suffix = Path(filename).suffix.lower()
    content_type = DOCUMENT_MIME_TYPES.get(suffix)
    if not content_type:
        raise ValueError("only TXT, CSV, XLSX, and DOCX files are supported")
    if not content or len(content) > MAX_DOCUMENT_BYTES:
        raise ValueError("document must be between 1 byte and 10 MB")
    try:
        if suffix == ".txt":
            text = content.decode("utf-8-sig")
        elif suffix == ".csv":
            text = _csv_text(content)
        elif suffix == ".xlsx":
            text = _xlsx_text(content)
        else:
            text = _docx_text(content)
    except (UnicodeDecodeError, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise ValueError("file could not be read as the declared document type") from exc
    text = _compact(text)
    if not text:
        raise ValueError("the document contains no readable text")
    if len(text) > MAX_EXTRACTED_CHARS:
        text = text[:MAX_EXTRACTED_CHARS] + "\n\n[本地文件内容已截取，原文件仍保存在本机]"
    return content_type, text


def _csv_text(content: bytes) -> str:
    rows = csv.reader(io.StringIO(content.decode("utf-8-sig")))
    lines: list[str] = []
    for index, row in enumerate(rows):
        if index >= MAX_SPREADSHEET_ROWS:
            lines.append("[CSV 已截取前 5000 行]")
            break
        lines.append(" | ".join(str(cell).strip() for cell in row[:MAX_SPREADSHEET_COLUMNS]))
    return "\n".join(lines)


def _xlsx_text(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("XLSX parsing dependency is unavailable") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines: list[str] = []
    remaining_rows = MAX_SPREADSHEET_ROWS
    try:
        for sheet in workbook.worksheets[:MAX_SHEETS]:
            if remaining_rows <= 0:
                lines.append("[Excel 已截取前 5000 行]")
                break
            lines.append(f"## 工作表：{sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                if remaining_rows <= 0:
                    lines.append("[Excel 已截取前 5000 行]")
                    break
                values = ["" if value is None else str(value).strip() for value in row[:MAX_SPREADSHEET_COLUMNS]]
                if any(values):
                    lines.append(" | ".join(values))
                    remaining_rows -= 1
    finally:
        workbook.close()
    return "\n".join(lines)


def _docx_text(content: bytes) -> str:
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    lines: list[str] = []
    for paragraph in root.iter(f"{namespace}p"):
        text = "".join(node.text or "" for node in paragraph.iter(f"{namespace}t")).strip()
        if text:
            lines.append(text)
    return "\n".join(lines)


def _compact(text: str) -> str:
    return re.sub(r"\n{3,}", "\n\n", text.replace("\x00", "")).strip()
